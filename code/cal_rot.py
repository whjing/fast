#%%
#!/usr/bin/env python
################################################################################
#  This is the code to calibrate polarization data observed with the FAST      #
#  telescope                                                                   # 
#      The code:                                                               #
#         - measures calibration parameters from the input calibration scans   #
#         - calibrate ALL rpf files in the current directory                   #
#         - bin the data in df-MHz wide bands                                  #
#         - write sdfits output file.                                          #
#                                                                              #
#      It assumes the following scans are available:                           #
#         - an unpolarised flux calibrator (1934, 0407, and HydA only for now) #
#         - a polarised calibrator (0043-424 and 3C138 only for now).          #
################################################################################


from astropy.io import fits as pf
from astropy.modeling import models, fitting
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import numpy as np
import astropy.units as u
from astropy.time import Time
from astropy.coordinates import SkyCoord, ITRS, EarthLocation, AltAz
from astropy.stats import sigma_clip
from scipy import interpolate
import csv
import sys
import os
import glob
import re as RegEx
#from kapteyn import maputils
from astropy import wcs
from scipy.interpolate import griddata
#from pyslalib import slalib
from scipy import ndimage
import gc
import sys
from natsort import natsorted
from scipy.signal import medfilt




# FAST locations
FAST_lat = 25.+39./60.+10.6/3600.
FAST_lon = 106.+51./60.+24.0/3600.
FAST_alt = 1110.0288
FAST_loc = EarthLocation(lat=FAST_lat*u.deg, lon=FAST_lon*u.deg, height=FAST_alt*u.m) 
UTC_OFFSET = 8.*u.hour
VC = 299792458. 
DIAM = 300.

#from astroplan import Observer
#from pytz import timezone
#observer = Observer(name='FAST Telescope',
#                    location = FAST_loc,
#                    timezone = timezone('Asia/Shanghai'))

# In case, finals2000A cannot be downloaded
from astropy.utils import iers
#iers.conf.auto_download = False
#iers.IERS.iers_table = iers.IERS_A.open("/home/xiaohui/erhai/FAST/finals2000A.all")
#iers.IERS.iers_table = iers.IERS_A.open("/home/xhsun/Processing/FAST/finals2000A.all")

from astropy.coordinates import Angle
BeamOffsetDict={}
BeamOffsetDict['M08']= [Angle(11.69*u.arcmin), 1.0, 1.0, Angle(0.*u.deg)] 
BeamOffsetDict['M02']= [Angle(5.89*u.arcmin), 1.0, 1.0, Angle(0.*u.deg)] 
BeamOffsetDict['M05']= [Angle(5.93*u.arcmin), -1.0, -1.0, Angle(0.*u.deg)] 
BeamOffsetDict['M14']= [Angle(11.78*u.arcmin), -1.0, -1.0, Angle(0.*u.deg)] 

BeamOffsetDict['M19']= [Angle(10.14*u.arcmin), 1.0, 1.0, Angle(30.*u.deg)] 
BeamOffsetDict['M13']= [Angle(10.17*u.arcmin), -1.0, -1.0, Angle(30.*u.deg)] 

BeamOffsetDict['M09']= [Angle(10.2*u.arcmin), 1.0, 1.0, Angle(-30.*u.deg)] 
BeamOffsetDict['M15']= [Angle(10.07*u.arcmin), -1.0, -1.0, Angle(-30.*u.deg)] 

BeamOffsetDict['M10']= [Angle(11.74*u.arcmin), 1.0, 1.0, Angle(-60.*u.deg)] 
BeamOffsetDict['M03']= [Angle(5.885*u.arcmin), 1.0, 1.0, Angle(-60.*u.deg)] 
BeamOffsetDict['M06']= [Angle(5.835*u.arcmin), -1.0, -1.0, Angle(-60.*u.deg)] 
BeamOffsetDict['M16']= [Angle(11.67*u.arcmin), -1.0, -1.0, Angle(-60.*u.deg)] 

BeamOffsetDict['M18']= [Angle(11.775*u.arcmin), 1.0, 1.0, Angle(60.*u.deg)] 
BeamOffsetDict['M07']= [Angle(5.92*u.arcmin), 1.0, 1.0, Angle(60.*u.deg)] 
BeamOffsetDict['M04']= [Angle(5.76*u.arcmin), -1.0, -1.0, Angle(60.*u.deg)] 
BeamOffsetDict['M12']= [Angle(11.645*u.arcmin), -1.0, -1.0, Angle(60.*u.deg)] 

BeamOffsetDict['M17']= [Angle(10.15*u.arcmin), -1.0, -1.0, Angle(-90.*u.deg)] 
BeamOffsetDict['M11']= [Angle(10.15*u.arcmin), 1.0, 1.0, Angle(-90.*u.deg)] 


def saveListFile(allList, fields, output):
    """
    save a list to a csv file
    """
    if len(fields) != len(allList):
        print("Size of allList and fields do not match!")
        sys.exit()

    allrows = []
    for i in range(allList[0].size):
        row = {}
        for j in range(len(fields)):
            row[fields[j]] = allList[j][i]
        allrows.append(row)

    csvfile = open(output, "w")
    obj = csv.DictWriter(csvfile, fieldnames=fields)
    obj.writeheader()
    obj.writerows(allrows)
    csvfile.close()


def readListFile(inFile, fields):
    """
    read a list from a csv file
    """
    csvfile = open(inFile, "r")
    obj = csv.DictReader(csvfile)
    fields_r = obj.fieldnames

    for term in fields:
        if term not in fields_r:
            print(term + 'not in fields!')
            sys.exit()
    
    allList = []
    for i in range(len(fields)):
        allList.append(np.array([0.]))

    num = 0
    for row in obj:
        for i in range(len(fields)):
            if num == 0:
                allList[i][num] = float(row[fields[i]])
            else:
                allList[i] = np.append(allList[i], float(row[fields[i]]))
        num += 1

    return allList


def flux_model(source, freq):
    """
    get flux density (Jy) of a calibrator at a given frequency (GHz) 
    The model is from Perley & Butler (2017)
    lg(S) = a0 + a1*lg(f) + a2*(lgf)**2 + a3*(lgf)**3 ...
    """
    sou = {}
    sou['3C48'] = [1.3253, -0.7553, -0.1914, 0.0498]
    sou['3C138'] = [1.0088, -0.4981, -0.1552, -0.0102, 0.0223]
    sou['3C286'] = [1.2481, -0.4507, -0.1798, 0.0357]
    sou['3C380'] = [1.1941, -0.5680, -0.1755, -0.0251, 0.0072]#very very uncertain(average 3)

    if source.upper() not in sou:
        print("Calibrator ", source.upper(), " not availble!")
        return None 
    else:
        value = sou[source.upper()]

        print(f'value:{value}')
        s = 0.
        lgf = np.log10(freq)
        for i in range(len(value)):
            s += value[i] * lgf ** float(i) 
    print(f's:{s}')
    return 10**s


def pol_model(source,freq):
    """
    get polarization percentage and polarization angle
    Based on Perley & Butler (2013)
    freq in GHz
    """
    if source.upper() == '3C286':
        #pc = 8.6 # 1.05 GHz
        #pc = 9.5 # 1.45 GHz
        pc = 9.1 # assume an average
        pa = 33.
    elif source.upper() == '3C138':
        #pc = 5.6 # 1.05 GHz
        #pc = 7.5 # 1.45 GHz
        pc = 6.6 # assume an average
        #pa = -11  # 1.45 GHz
        #pa = -14  # 1.05 GHz
        pa = -12.5 # assume an average
    elif source.upper() == '3C48':
        #pc = 0.3 # 1.05 GHz
        #pc = 0.5 # 1.45 GHz
        pc = 0.4 # assume an average
        #pa = 25 # 1.05 GHz
        #pa = 140 # 1.45 GHz
        pa = 25 # very uncertain
    elif source.upper() == '3C380':# very very uncertain(average 3)
        pc = 5.1
        pa = 15

    return pc, pa


def cal_test(dir_path, file_out,average=64):
    # 一列数据
    file_list = natsorted(glob('{work_dir}/high_202410/T*.dat'))
    data_all = np.fromfile("{work_dir}/high_202410/freq.dat",dtype=float, sep=" ")
    # average 64
    data_all = data_all.reshape(-1,64)
    data_all = np.nanmedian(data_all,axis=1)
    print(data_all.shape)
    # data_all = freq0 +  np.linspace(0., 65535., int(65536/average)) * df
    for file in file_list:
        data = np.fromfile(file,dtype=float, sep=" ")
        print(file)
        print(data.shape)
        # average 64
        data = data.reshape(-1,64)
        print(data.shape)
        # average 64
        data = np.nanmedian(data,axis=1)
        print(data.shape)
        plt.plot(data)
        data_all = np.vstack((data_all,data))

    print(data_all.shape)
    np.savetxt(file_out,data_all.T,fmt='%f')


class Track:

    def __init__(self, posFile, kyFile=None, col_time='A', col_SDP_X='H', col_SDP_Y='I', col_SDP_Z='J'):
        """
        read in the excel file containing two sheets
        SDP_PhaPos_X, SDP_PhaPos_Y, SDP_PhasPos_Z will be used to calculate alt and az
        The corresponding columns are: H, I, and J.
        Time is in local time (Beijing time UTC+8) and the column is A. 
        
        Or read directly all the position information from posfile
        """
        self.output = posFile
        self.fields = ['MJD', 'RA2000', 'DEC2000', 'AZ', 'ALT']
        if kyFile is not None: 
            wb = load_workbook(kyFile)
            sheet = wb.worksheets[0]
            Nrows = sheet.max_row
            timeList = []
            mjd_sdp = np.zeros(Nrows)
            x_sdp = np.zeros(Nrows)
            y_sdp = np.zeros(Nrows)
            z_sdp = np.zeros(Nrows)
            nn = 0
            for row in range(2, Nrows+1):
                t = Time(sheet[col_time+str(row)].value.encode('ascii')) - UTC_OFFSET
                x = sheet[col_SDP_X+str(row)].value
                y = sheet[col_SDP_Y+str(row)].value
                z = sheet[col_SDP_Z+str(row)].value
                # exclude NaNs and 0 for x, y, z.
                if np.isnan(x) or np.isnan(y) or np.isnan(z) or x==0 or y==0 or z==0: continue
                mjd_sdp[nn] = t.mjd  
                x_sdp[nn] = x
                y_sdp[nn] = y
                z_sdp[nn] = z
                timeList.append(t)
                nn += 1

            self.t = timeList
            self.mjd = mjd_sdp[:nn]
            self.x = x_sdp[:nn]
            self.y = y_sdp[:nn]
            self.z = z_sdp[:nn]
            self.pressure = 0.
            self.temperature = 0.
        else:
            [self.mjd, self.ra, self.dec, self.az, self.alt] = readListFile(posFile, self.fields)
             

    def xyz2azalt(self):
        """
        convert x, y, z to az, alt
        Return: az and alt in degrees
        """
        R_ca = np.sqrt(self.x*self.x+self.y*self.y+self.z*self.z)
        az0 = np.degrees(np.arctan2(self.x, self.y)) % 360.
        alt = 90. - np.degrees(np.arccos(-self.z/R_ca))
        az = np.copy(az0)

        # +180 if in the range [0,180.]
        sel1 = az0 >= 0.
        sel2 = az0 <= 180.
        sel = sel1 * sel2
        az[sel] += 180.

        # -180 if in the range (180, 360]
        sel1 = az0 > 180.
        sel2 = az0 <= 360.
        sel = sel1 * sel2
        az[sel] -= 180.

        self.az = az
        self.alt = alt


    def azalt2eq(self):
        """
        Convert az, alt to ra, dec
        Return: ra and dec in degrees
        """
        frame = AltAz(obstime=self.t, location=FAST_loc, pressure=self.pressure, temperature=self.temperature)
        c = SkyCoord(alt = self.alt*u.deg, az = self.az*u.deg, frame = frame)
        c_new = c.transform_to('fk5')
        self.ra = c_new.ra.deg
        self.dec = c_new.dec.deg


    def plotRADec(self,output=None):
        fig = plt.figure()
        ax = fig.add_subplot(1,1,1)
        ax.plot(self.ra, self.dec) 
        ax.set_xlabel('RA')
        ax.set_ylabel('Dec')
        if output==None: plt.show()
        else: plt.savefig(output)


    def runsave(self):
        """
        output as csv file
        """
        self.xyz2azalt()
        self.azalt2eq()
        allList = [self.mjd, self.ra, self.dec, self.az, self.alt]
        saveListFile(allList, self.fields, self.output)

class Scan:
    def __init__(self, ObsFile, fileList, t_track, ra_track, dec_track, RA_c, Dec_c, scanPts, BeamNr, RotAngle, Chan=42000):

        # read in all the data
        print("Reading all the data ....")
        print("ObsFile:",ObsFile)   
        data = pf.getdata(ObsFile)
        hdu = pf.open(ObsFile)[1]
        hdr = hdu.header
        

        self.ObsTime = data['UTOBS'] 
        self.ObsData = data['DATA'][:,Chan,:]
        # 将obsdata中的nan替换为0
        self.ObsData[np.isnan(self.ObsData)] = 0
        self.freq0 = hdr['FREQ0'] + hdr['CHAN_BW']/2
        self.bw = hdr['CHAN_BW']
        self.fghz = (self.freq0 + Chan * self.bw)/1.e3

        shape = self.ObsData.shape
        self.NumStokes = shape[1]

        print("Interpolation ...")
        f1 = interpolate.interp1d(t_track, ra_track, kind='linear')
        f2 = interpolate.interp1d(t_track, dec_track, kind='linear')

        self.ra = f1(self.ObsTime)
        self.dec = f2(self.ObsTime)

        self.raList = []
        self.decList = []
        self.dataList = []
        self.distList = []

        self.nscan = len(scanPts)
        
        for i in range(self.nscan):#self.nscan
            # print(self.ObsTime[0],scanPts[i][0],scanPts[i][1],self.ObsTime[-1])
            # print("scanPts[i][0]:",scanPts[i][1]-self.ObsTime[])
            sel1 = self.ObsTime >= scanPts[i][0]
            sel2 = self.ObsTime <= scanPts[i][1]
            sel = sel1 & sel2

            # print("sel1:",self.ObsData[sel1])
            # print("sel2:",self.ObsData[sel2])
            #print("sel:",sel)
            # print("self.ObsData:",self.ObsData[sel])
            a=0
            b=0
            for k in range(len(sel)):
                if sel[k]==False:
                    a+=1
                else:
                    b+=1
            # print(i)
            # print("a:",a)
            # print("b:",b)

            # print(self.ObsData[sel])

            self.raList.append(self.ra[sel])
            self.decList.append(self.dec[sel])
            self.dataList.append(self.ObsData[sel])


        if int(BeamNr[1:]) > 1:
            ss = BeamNr
            for i in range(self.nscan):
                dec_t = np.radians(self.decList[i])
                ra_t = np.radians(self.raList[i])
                dec_t = dec_t + BeamOffsetDict[ss][1] * BeamOffsetDict[ss][0].rad * np.sin((RotAngle + BeamOffsetDict[ss][3]).rad)
                ra_t = ra_t + BeamOffsetDict[ss][2] * BeamOffsetDict[ss][0].rad * np.cos((RotAngle + BeamOffsetDict[ss][3]).rad)/np.cos(dec_t)
                self.decList[i] = np.degrees(dec_t)
                self.raList[i] = np.degrees(ra_t)

        c0 = SkyCoord(ra=RA_c*u.deg, dec=Dec_c*u.deg)
        for i in range(self.nscan):
            c = SkyCoord(ra=self.raList[i]*u.deg, dec=self.decList[i]*u.deg)
            self.distList.append(c0.separation(c).arcmin)

        
    def CalWithInjectedNoise(self, Tcal=[10., 10.], check=False, signQ=1.0, signU=1.0, RotAngle=Angle(0.*u.deg)):
        # separate Sig and Cal
        N_r_p = 8 
        N_c_p = 2

        self.I_all = []
        self.I_ave = []
        self.I_sig = []
        self.Q_all = []
        self.Q_ave = []
        self.Q_sig = []
        self.U_all = []
        self.U_ave = []
        self.U_sig = []
        self.V_all = []
        self.V_ave = []
        self.V_sig = []
        self.dist_all = []
        self.dist_ave = []
        self.dist_sig = []
        # print("self.nscan:",self.nscan)
 
        for jj in range(self.nscan):  

            if check: fig = plt.figure()

            sig_all = []
            ref_all = []
            sign = 1.
            for i in range(self.NumStokes): 
                # print(f"i:{i}")
                if check:
                    ax1 = fig.add_subplot(N_r_p,N_c_p,i+1)
                    ax2 = fig.add_subplot(N_r_p,N_c_p,i+5)
                Toff = self.dataList[jj][::2,i]
                Ton  = self.dataList[jj][1::2,i]
                # print("self.dataList[0]:",self.dataList[0])
                # print("Toff:",Toff)
                # print("Ton:",Ton)

                if i==0: nsize = min(Toff.size, Ton.size)
                sig = Ton[:nsize] + Toff[:nsize]
                ref = Ton[:nsize] - Toff[:nsize]
                if i==0 and ref[ref>0].size - ref[ref<0].size < 0: sign = -1.

                ref *= sign

                # linear fit of ref signal
                #p1 = models.Polynomial1D(1)
                #pfit = fitting.LinearLSQFitter()
                #sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                #x = np.arange(ref.size)
                #new_model, mask = sigma_clip_fit(p1, x, ref)
                #ref_fit = new_model(x)

                ref_fit = np.median(ref)

                if check:
                    #ax1.plot(x, ref)
                    #ax1.plot(x, ref_fit)
                    #ax2.plot(x, sig)
                    ax1.plot(ref)
                    ax1.axhline(ref_fit)
                    ax2.plot(sig)

                sig_all.append((sig - ref_fit) / 2.)
                ref_all.append(ref_fit * 1.)

            #   I' = | 1        scale_IQ |  I
            #   Q' = | scale_IQ    1     |  Q
            I_ref = (ref_all[0] + ref_all[1])/2.
            Q_ref = (ref_all[0] - ref_all[1])/2.
            U_ref = ref_all[2]*1.
            V_ref = ref_all[3]*1.
            scale_IQ = Q_ref / I_ref 
            # print(f"Q_ref:{Q_ref}", f"I_ref:{I_ref}")
            ang_inst = np.arctan2(V_ref, U_ref) 
        
            I_ref_c = (I_ref - scale_IQ * Q_ref) / (1 - scale_IQ*scale_IQ)
            Q_ref_c = (Q_ref - scale_IQ * I_ref) / (1 - scale_IQ*scale_IQ)
            U_ref_c = U_ref * np.cos(ang_inst) + V_ref * np.sin(ang_inst)
            V_ref_c = V_ref * np.cos(ang_inst) - U_ref * np.sin(ang_inst)

            f_U = I_ref_c / U_ref_c
  
            scale_T = np.sqrt(Tcal[0]*Tcal[1]) / np.sqrt(ref_all[0]*ref_all[1]) 

            I_ref_c *= scale_T
            Q_ref_c *= scale_T
            U_ref_c *= scale_T
            V_ref_c *= scale_T

            I_sig = (sig_all[0] + sig_all[1])/2.
            Q_sig = (sig_all[0] - sig_all[1])/2.
            U_sig = sig_all[2]*1.
            V_sig = sig_all[3]*1.

            I_sig_c = (I_sig - scale_IQ * Q_sig) / (1 - scale_IQ*scale_IQ)
            Q_sig_c = (Q_sig - scale_IQ * I_sig) / (1 - scale_IQ*scale_IQ)
            U_sig_c = U_sig * np.cos(ang_inst) + V_sig * np.sin(ang_inst)
            V_sig_c = V_sig * np.cos(ang_inst) - U_sig * np.sin(ang_inst)

            U_sig_c *= f_U
  
            I_sig_c *= scale_T
            Q_sig_c = Q_sig_c * scale_T * signQ
            U_sig_c = U_sig_c * scale_T * signU
            V_sig_c *= scale_T

            # correction for feed rotation
            Q_sig_c1 = Q_sig_c * np.cos(2.*RotAngle).value + U_sig_c * np.sin(2.*RotAngle).value
            U_sig_c1 = -Q_sig_c * np.sin(2.*RotAngle).value + U_sig_c * np.cos(2.*RotAngle).value
            Q_sig_c = Q_sig_c1 * 1.
            U_sig_c = U_sig_c1 * 1.

            ttt = (self.distList[jj][::2][:nsize]+self.distList[jj][1::2][:nsize])/2.
            #print("jj:",jj)
            #print("mean(ttt):",np.mean(ttt))
            self.dist_all.append(ttt)
            self.dist_ave.append(np.mean(ttt))
            # print("self.dist_ave:",self.dist_ave)
            self.dist_sig.append(np.std(ttt))

            self.I_all.append(I_sig_c)
            self.Q_all.append(Q_sig_c)
            self.U_all.append(U_sig_c)
            self.V_all.append(V_sig_c)

            self.I_ave.append(np.mean(I_sig_c))
            self.I_sig.append(np.std(I_sig_c))
            self.Q_ave.append(np.mean(Q_sig_c))
            self.Q_sig.append(np.std(Q_sig_c))
            self.U_ave.append(np.mean(U_sig_c))
            self.U_sig.append(np.std(U_sig_c))
            self.V_ave.append(np.mean(V_sig_c))
            self.V_sig.append(np.std(V_sig_c))

            if check:
                #ax9 = fig.add_subplot(N_r_p, N_c_p, 9)
                #ax9.plot(I_ref_c)
                #ax10 = fig.add_subplot(N_r_p, N_c_p, 10)
                #ax10.plot(Q_ref_c)
                #ax11 = fig.add_subplot(N_r_p, N_c_p, 11)
                #ax11.plot(x, scale_IQ)
                #ax11.plot(U_ref_c)
                #ax12 = fig.add_subplot(N_r_p, N_c_p, 12)
                #ax12.plot(x, ang_inst)
                #ax12.plot(V_ref_c)
                ax13 = fig.add_subplot(N_r_p, N_c_p, 13)
                ax13.plot(I_sig_c)
                ax14 = fig.add_subplot(N_r_p, N_c_p, 14)
                ax14.plot(Q_sig_c)
                ax15 = fig.add_subplot(N_r_p, N_c_p, 15)
                ax15.plot(U_sig_c)
                ax16 = fig.add_subplot(N_r_p, N_c_p, 16)
                ax16.plot(V_sig_c)
                plt.show()
        #print("self.dist_ave:",self.dist_ave[0])


    def GetPeakValues(self, x=None, x_sig = None, signals=None, signals_sig = None, check=False, figname=None, output=None, thres=10.):
        # get the gain factor by ffitting calibrator
     
        if signals == None:
            signals = [np.array(self.I_ave), np.array(self.Q_ave), np.array(self.U_ave), np.array(self.V_ave)]
            signals_sig = [np.array(self.I_sig), np.array(self.Q_sig), np.array(self.U_sig), np.array(self.V_sig)]

        if x == None:
            x = np.array(self.dist_ave)
            #print("x:",x)
            x_sig = np.array(self.dist_sig)

        if check:
            fig = plt.figure()
            ax = fig.add_subplot(4,1,1)
            #for i in range(len(self.I_all)):
            for i in range(1):
                ax.plot(np.array(self.dist_all[i]), np.array(self.I_all[i]), 'o')

            ax = fig.add_subplot(4,1,2)
            #for i in range(len(self.Q_all)):
            for i in range(1):
                ax.plot(np.array(self.dist_all[i]), np.array(self.Q_all[i]), 'o')

            ax = fig.add_subplot(4,1,3)
            #for i in range(len(self.U_all)):
            for i in range(1):
                ax.plot(np.array(self.dist_all[i]), np.array(self.U_all[i]), 'o')

            ax = fig.add_subplot(4,1,4)
            #for i in range(len(self.V_all)):
            for i in range(1):
                ax.plot(np.array(self.dist_all[i]), np.array(self.V_all[i]), 'o')

            plt.show()


        if output is not None:
            f = open(output, 'w')
            for i in range(x.size):
                ss = str(x[i])+ '  ' + str(x_sig[i]) + ' ' + str(signals[0][i])+ ' ' + str(signals_sig[0][i]) + ' ' + str(signals[1][i])+ ' ' + str(signals_sig[1][i]) + ' ' + str(signals[2][i])+ ' ' + str(signals_sig[2][i]) + ' ' + str(signals[3][i])+ ' ' + str(signals_sig[3][i]) + '\n'
                f.write(ss)
            f.close()

        self.sig_obs = []

        for i in range(len(signals)):
            sig = signals[i]

            if i==0:

                x_f = x
                y_f = sig

                """
                y_0 = y_f[np.argmax(x_f)]
                p_init = models.Polynomial1D(1)
                p_init.c0 = y_0
                ytt = y_f - y_0
                mean = np.min(x_f)
                amp = ytt[np.argmin(x_f)]
                std = np.sqrt(np.sum((x_f-mean)*(x_f-mean)*ytt)/np.sum(ytt))
                g_init = models.Gaussian1D(amplitude=amp, mean=mean, stddev=std)
                g_init.mean.fixed = True
                pg = p_init + g_init
                fit = fitting.LevMarLSQFitter()
                fitted_model = fit(pg, x_f, y_f)
                ierr = fit.fit_info['ierr']
                Intensity = fitted_model[1].amplitude.value
                beam = fitted_model[1].stddev.value*2.*np.sqrt(2.*np.log(2.))
                self.sig_obs.append([Intensity, beam, mean])
                """
                
                Intensity = y_f[np.argmin(x_f)] - np.mean(y_f[x_f>thres])
                # print(np.argmin(x_f))
                # print(np.mean(y_f[x_f>thres]))
                #print("Intensity:",Intensity)
                self.sig_obs.append([Intensity])

                if check == True:
                    if i==0: fig = plt.figure(figsize=(10,8))
                    ax1 = fig.add_subplot(3,2,i+1)
                    ax1.plot(x_f, y_f, 'o')
                    ax1.errorbar(x_f, y_f, xerr=x_sig, yerr=signals_sig[i], marker='+', ls='none')
                    ax1.axhline(y_f[np.argmin(x_f)])
                    ax1.axhline(np.mean(y_f[x_f>thres]))
                    #ax1.plot(x_tmp, fitted_model(x_tmp))
                    #ax2 = fig.add_subplot(3,2,i+5)
                    #tttt = y_f - fitted_model(x_f)
                    #ax2.plot(x_f, tttt)
            else:

                self.sig_obs.append([sig[np.argmin(x)] - np.mean(sig[x>thres])])
            
                if check == True:
                    ax1 = fig.add_subplot(3,2,i+1)
                    ax1.plot(x, sig, 'o')
                    ax1.errorbar(x, sig, xerr=x_sig, yerr=signals_sig[i], marker='+', ls='none')
                    ax1.axhline(sig[np.argmin(x)])
                    ax1.axhline(np.mean(sig[x>thres]))

        ti = self.sig_obs[0][0]
        q = self.sig_obs[1][0]
        u = self.sig_obs[2][0]
        v = self.sig_obs[3][0]
        pi = np.sqrt(q*q + u*u)
        pc = pi*100./ti
        pa = np.degrees(0.5*np.arctan2(u, q))

        if check == True:
            if figname == None: plt.show()
            else: 
                plt.savefig(figname)
                plt.cla()
                plt.clf()
                plt.close('all')
                plt.close(fig)
                gc.collect()

        return ti, q, u, v, pi, pa, pc



    def GetCalFactors_S(self, source='3C138', freq = None, signals = None):#source='3C286'
        #print('get')

        if signals == None:
            signals = []
            for i in range(len(self.sig_obs)):
                #print(self.sig_obs[i][0])
                signals.append(self.sig_obs[i][0])

        if freq == None: freq = self.fghz
        print(f'---I Q U--')

        I0 = flux_model(source, freq)
        # print(f"I0:{I0}")
        pc, pa = pol_model(source, freq)
        Q0 = I0*pc*np.cos(2.*np.radians(pa))/100.
        # print(f"Q0:{Q0}")
        U0 = I0*pc*np.sin(2.*np.radians(pa))/100.
        # print(f"U0:{U0}")
        V0 = 0.
        # print(f"signals:{signals}")

        f_I = I0/ signals[0]
        f_Q = Q0 / signals[1]
        f_U = U0 / signals[2]
        # print(f'FI,FQ,FU: {f_I},{f_Q},{f_U}')

        # print(I0)
        # print(pc)
        # print(signals[0])

        return f_I, f_Q, f_U


def makeCalScan(posFile, scan_pts_file, infTemplate, fileList, BeamList, RA_c, Dec_c, source, Nchan_N, good_chan_file = None, tcalFile=None, RotAngle=Angle(0.*u.deg), ChanList=[42000], signQ=-1.0, signU=1.0, check=False, outFileBase='all'):
    print(f'make scan {fileList}')
    track = Track(posFile)
    

    beamScanPts = [] 
    f = open(scan_pts_file, 'r')#打开按观测到3C138先后时间排序好的beam顺序文件
    allLines = f.readlines()
    f.close()
    for term in allLines:
        x = term.split()
        beamScanPts.append([float(x[1]), float(x[2])])#存储上诉排序

    if tcalFile != None:
        print(tcalFile)
        tcals0 = np.fromfile(tcalFile, sep="\n")
        tcals = tcals0.reshape((tcals0.size//39, 39))

    if good_chan_file != None:
        print(good_chan_file)
        chan_all = np.loadtxt(good_chan_file, dtype=int)
        # 如果该索引的平均值小于0.5，说明该索引通道是好通道
        good_chans = np.where(chan_all.mean(axis=0) > 0.6)[0]
        
    print(ChanList)

    for Chan in ChanList: 
        print("Channel: ", Chan)
        if good_chan_file != None:
            # 如果Chan在good_chans中，说明是好通道
            if Chan in good_chans: 
                print(f"++++++ Bad chans {Chan} +++++++++++")
                continue

        for i in BeamList:
            print(f"Channel: {Chan} Beam: {i}")
            BeamNr = "M%02d" % i

            inf = infTemplate.replace('YYY', BeamNr[1:])
            print(f'inf{inf}')
            obs = Scan(inf, fileList, track.mjd, track.ra, track.dec, RA_c, Dec_c, beamScanPts, BeamNr, RotAngle, Chan=Chan)
            if tcalFile != None:
                freq_t = tcals[:,0]
                tcal_xx = tcals[:,2*i-1]
                tcal_yy = tcals[:,2*i]
                freq0 = obs.fghz*1.e3
                tcal_interp_xx = interpolate.interp1d(freq_t, tcal_xx, kind='linear')
                tcal_interp_yy = interpolate.interp1d(freq_t, tcal_yy, kind='linear')
                Tcal = [tcal_interp_xx(freq0), tcal_interp_yy(freq0)]
                # print(111)
            else:
                Tcal = [10., 10.]
            obs.CalWithInjectedNoise(Tcal = Tcal, check=check, signQ=signQ, signU=signU, RotAngle=RotAngle)
            ti, q, u, v, pi, pa, pc = obs.GetPeakValues()
            f_I, f_Q, f_U = obs.GetCalFactors_S(source=source)
            # print(f_I)
            # sys.exit()
            print(f'f_I: {f_I}')
            print(f'outFileBase: {outFileBase}')
            outfile = outFileBase
            print(f'outfile {outfile}')
            fout = open(outfile, 'a+')
            ss = str(i)+" "+str(Chan)+" "+str(f_I)+" "+str(f_Q)+" "+str(f_U)+" "+str(ti)+" "+str(q)+" "+str(u)+" "+str(v)+" "+str(pi)+" "+str(pa)+" "+str(pc)+" "+str(obs.fghz)+"\n"
            fout.write(ss)
            fout.close()


def obtain_gainratio(file_in, file_out, plot=True):


    Ncol = 13
    data1 = np.loadtxt(file_in)#, sep='\n')
    print(f'orignal data shape {data1.shape}')

    freq = data1[:,-1][::19]
    print(freq[1])

    print(f'freq data shape {freq.shape}')
    gain1 = 1./data1[:,2]
    print(f'gain1 data shape {gain1.shape}')
    N_beams = 19

    ouf = open(file_out, 'w')
    ouf.write('1, 0\n')

    y0 = gain1[0::19]
    print(f'y0 shape {y0.shape}')

    fig = plt.figure(figsize = (10,10))


    for beam in range(1, N_beams):
        beam_name = beam+1
        y1 = gain1[beam::19]
        # y = y0 / y1 
        y = y1 / y0 
        

        p = models.Polynomial1D(1)
        pfit = fitting.LinearLSQFitter()

        
        model = pfit(p, freq, y)
        yf = model(freq)
        ss = '%.2f, %.2f' % (model.c0.value, model.c1.value)
        ouf.write(ss+'\n')

        ymedfit = medfilt(y, kernel_size= 99)

        ax = fig.add_subplot(6,3,beam)
        ax.plot(freq, y, '.', color = 'lightblue')
        ax.plot(freq, yf, label = f'{ss}')
        ax.plot(freq, ymedfit, color = 'magenta', alpha = 0.5, label = 'medfit')
        ax.set_title(f'Beam {beam_name}')
        ax.set_xlabel('Frequency (GHz)')
        ax.set_ylabel(r'$G_{B}/G_{B1}$')
        plt.tight_layout()
        plt.legend()
        plt.savefig(f"{file_out}.png", dpi=100)


def main():

    posFile = f"../data_ave-64/3C138_20241207/3C138_2024_12_07_02_27.csv"
    infTemplate = f"../data_ave-64/3C138_20241207/3C138_multibeamcalibration-MYYY_ave-64.fits"
    scan_pts_file = f"../data_ave-64/3C138_20241207/3C138_2024_12_07_02_27_scan_pts.dat"
    outFileBase = "../data_ave-64/3C138_20241207/cat.dat"
    file_gainratio = f"../data_ave-64/3C138_20241207/gain_ratio.dat"
    source = "3C138"
    c0 = SkyCoord.from_name(source)
    RA_c, Dec_c = c0.ra.deg, c0.dec.deg
    BeamList = range(1,20)

    ChanList = np.arange(0,1023)
    signQ = -1.
    signU = 1.

    tcalDir = '../data_orig/noise'
    tcalFile = f'{tcalDir}/high_ave64.dat'
    good_chan_file = f"../data_ave-64/3C138_20241207/good_chans.txt"
    ang = 0. 
    check = False
    fileList=[0,1]
    Nchan_N =1
    if os.path.exists(outFileBase): os.remove(outFileBase)

    makeCalScan(posFile, scan_pts_file, infTemplate, fileList, BeamList, RA_c, Dec_c, source, good_chan_file = good_chan_file, tcalFile=tcalFile, RotAngle=Angle(ang*u.deg), ChanList=ChanList, signQ=signQ, signU=signU, check=check, outFileBase=outFileBase, Nchan_N=Nchan_N)

    obtain_gainratio(outFileBase, file_gainratio, plot=True)

if __name__ == "__main__":
    main()
