#%%
from glob import glob
import numpy as np
from astropy.io import fits as pf
from natsort import natsorted
import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import medfilt
from scipy import sparse
from scipy.sparse.linalg import spsolve
import os
import sys
from astropy.modeling import models, fitting
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import numpy as np
import astropy.units as u
from astropy.time import Time
from astropy.coordinates import SkyCoord, ITRS, EarthLocation, AltAz
from astropy.stats import sigma_clip
from scipy import interpolate
import csv
import sys
import os
import re as RegEx
# from kapteyn import maputils
from astropy import wcs
from scipy.interpolate import griddata
# from pyslalib import slalib
from scipy import ndimage
import gc

# FAST locations
FAST_lat = 25.+39./60.+10.6/3600.
FAST_lon = 106.+51./60.+24.0/3600.
FAST_alt = 1110.0288
FAST_loc = EarthLocation(lat=FAST_lat*u.deg, lon=FAST_lon*u.deg, height=FAST_alt*u.m) 
UTC_OFFSET = 8.*u.hour
VC = 299792458. 
DIAM = 300.


from astropy.utils import iers
#iers.conf.auto_download = False
#iers.IERS.iers_table = iers.IERS_A.open("/home/xiaohui/erhai/FAST/finals2000A.all")
#iers.IERS.iers_table = iers.IERS_A.open("/home/xhsun/Processing/FAST/finals2000A.all")

from astropy.coordinates import Angle
BeamOffsetDict={}
BeamOffsetDict['M08']= [Angle(11.69*u.arcmin), 1.0, 1.0, Angle(0.*u.deg)] 
BeamOffsetDict['M02']= [Angle(5.89*u.arcmin), 1.0, 1.0, Angle(0.*u.deg)] 
BeamOffsetDict['M05']= [Angle(5.93*u.arcmin), -1.0, -1.0, Angle(0.*u.deg)] 
BeamOffsetDict['M14']= [Angle(11.78*u.arcmin), -1.0, -1.0, Angle(0.*u.deg)] 

BeamOffsetDict['M19']= [Angle(10.14*u.arcmin), 1.0, 1.0, Angle(30.*u.deg)] 
BeamOffsetDict['M13']= [Angle(10.17*u.arcmin), -1.0, -1.0, Angle(30.*u.deg)] 

BeamOffsetDict['M09']= [Angle(10.2*u.arcmin), 1.0, 1.0, Angle(-30.*u.deg)] 
BeamOffsetDict['M15']= [Angle(10.07*u.arcmin), -1.0, -1.0, Angle(-30.*u.deg)] 

BeamOffsetDict['M10']= [Angle(11.74*u.arcmin), 1.0, 1.0, Angle(-60.*u.deg)] 
BeamOffsetDict['M03']= [Angle(5.885*u.arcmin), 1.0, 1.0, Angle(-60.*u.deg)] 
BeamOffsetDict['M06']= [Angle(5.835*u.arcmin), -1.0, -1.0, Angle(-60.*u.deg)] 
BeamOffsetDict['M16']= [Angle(11.67*u.arcmin), -1.0, -1.0, Angle(-60.*u.deg)] 

BeamOffsetDict['M18']= [Angle(11.775*u.arcmin), 1.0, 1.0, Angle(60.*u.deg)] 
BeamOffsetDict['M07']= [Angle(5.92*u.arcmin), 1.0, 1.0, Angle(60.*u.deg)] 
BeamOffsetDict['M04']= [Angle(5.76*u.arcmin), -1.0, -1.0, Angle(60.*u.deg)] 
BeamOffsetDict['M12']= [Angle(11.645*u.arcmin), -1.0, -1.0, Angle(60.*u.deg)] 

BeamOffsetDict['M17']= [Angle(10.15*u.arcmin), -1.0, -1.0, Angle(-90.*u.deg)] 
BeamOffsetDict['M11']= [Angle(10.15*u.arcmin), 1.0, 1.0, Angle(-90.*u.deg)]




def load_data(file_in, orignal=False, avetime=False, avechan=False):
    """读取 FITS 文件数据，并根据选项返回原始数据或进行时间/频率平均"""
    
    # 确保 avetime 和 avechan 不能同时为 True
    if sum([orignal, avetime, avechan]) != 1:
        raise ValueError("请确保 'orignal'、'avetime' 和 'avechan' 三者选择一个！")
    print(f'Loading: {file_in}')

    data = pf.getdata(file_in)
    obsData = data['DATA']

    # **原始数据**
    if orignal:
        XX = obsData[:, :, 0]
        YY = obsData[:, :, 1]
        XY = obsData[:, :, 2]
        YX = obsData[:, :, 3]

    # **时间平均**
    elif avetime:
        dataPath = file_in.replace('.fits', '_avetime.dat')
        if os.path.exists(dataPath):
            obsData = np.loadtxt(dataPath)
            print(f'Loading: {dataPath}')
        else:
            obsData = np.median(obsData, axis=0)  # 沿时间轴取中值
            np.savetxt(dataPath, obsData)
            print(f'{dataPath} has been saved!')

        XX, YY, XY, YX = obsData[:, 0], obsData[:, 1], obsData[:, 2], obsData[:, 3]

    # **频率平均**
    elif avechan:
        dataPath = file_in.replace('.fits', '_avechan.dat')
        if os.path.exists(dataPath):
            obsData = np.loadtxt(dataPath)
            print(f'Loading: {dataPath}')
        else:
            obsData = np.median(obsData, axis=1)  # 沿频率轴取中值
            np.savetxt(dataPath, obsData)
            print(f'{dataPath} has been saved!')

        XX, YY, XY, YX = obsData[:, 0], obsData[:, 1], obsData[:, 2], obsData[:, 3]

    I = (XX + YY)/2
    Q = (XX - YY)/2

    return I, Q



def find_obs_time(fitsFile, figName = None):
    # 修改文件路径
    obsAmpAll, ObsQc = load_data(fitsFile, avechan=True)
    print(f"All observation time is {obsAmpAll.shape[0]}")
    time = np.arange(obsAmpAll.shape[0])
    tAmp = np.column_stack((time, obsAmpAll))
    # select on data
    tAmpOn = tAmp[::2]
    timeOn = tAmpOn[:,0]
    print(f"tAMPOn shape is {tAmpOn.shape}")
    ampOn = tAmpOn[:,1]
    timeListNum = 1e4
    cutoff = np.max(ampOn)/3
    while timeListNum > 600:
        cutoff = cutoff * 1.1
        print(f"Now the cutoff is {cutoff}")
        timeList = tAmpOn[np.where(ampOn > cutoff)][:, 0]
        timeListNum = timeList.shape[0]
        print(f"Now the minimal timeListNum is {timeListNum }")
    
    timeStart = np.nanmin(timeList)
    timeEnd = np.nanmax(timeList)
    print(f"orignal time file: {timeStart} {timeEnd}")

    timeStart = max(0, timeStart)
    timeEnd = min(timeOn[-1], timeEnd)
    timeStart = int(timeStart)
    timeEnd = int(timeEnd)

    print(f"extend time file: {timeStart} {timeEnd}")

    

    recordText = [timeStart,timeEnd]
    print(recordText)
    return obsAmpAll, recordText



def process_data(file_in, column_index, kernel_size=1, iterations=1):

    I, Q  = load_data(file_in, avetime=True)
    num_chan_orig = I.shape[0]
    chans = np.arange(num_chan_orig)
    amp = I
    print(chans.shape, amp.shape)

    # 去除包含NaN的行
    valid_mask = ~np.isnan(chans) & ~np.isnan(amp)
    chans_valid = chans[valid_mask]
    amp_valid = amp[valid_mask]

    all_outliers = []

    fig, axes = plt.subplots(2, 1, figsize=(10, 6))

    for i in range(iterations):
        # 使用中值滤波
        kernel_size_in = kernel_size * (i+1)
        if kernel_size_in%2 == 0:
            kernel_size_in += 1
        print(kernel_size_in)
        y_medfilt = medfilt(amp_valid, kernel_size= kernel_size_in) #试过ArPLS，效果差不多
        residuals = amp_valid - y_medfilt
        filter_threshold = np.nanmean(y_medfilt) * 0.1/i

        # 识别超出阈值的值
        outliers = np.abs(residuals) > filter_threshold
        # 记录异常值
        outlier_data = np.column_stack((chans_valid[outliers], amp_valid[outliers]))
        
        num_chans_left = len(amp_valid) - outliers.sum()

        chans_valid = chans_valid[~outliers]
        amp_valid = amp_valid[~outliers]
        y_medfilt_plot = medfilt(amp_valid, kernel_size= kernel_size_in)

        if i == 0:
            ax0 = axes[0]
            ax0.scatter(chans_valid, amp_valid, marker='.', label='Data', s=1)
            ax0.plot(chans_valid, y_medfilt_plot, color='red')
            ax0.set_xlabel('Freq')
            ax0.set_ylabel('Intensity')
            ax0.set_ylim(np.min(amp_valid) *0.9, np.max(amp_valid) * 1.1)
            ax0.set_title('Before Flagging')
        elif i == iterations-1:
            ax1 = axes[1]
            ax1.scatter(chans_valid, amp_valid, marker='.', label='Data', s=1)
            ax1.plot(chans_valid, y_medfilt_plot, color='red')
            ax1.set_xlabel('Freq')
            ax1.set_ylabel('Intensity')
            ax1.set_ylim(np.min(amp_valid) *0.9, np.max(amp_valid) * 1.1)
            ax1.set_title(f"After Flagging, Iteration {i+1}, left {num_chans_left} chans")
            plt.savefig(f"{file_in.replace('.fits', f'_flaged.png')}")
        if num_chans_left < num_chan_orig/2:
            print(f'FLAG TOO MUCH BAD CHANS, STOP!')
            break
        else:
            all_outliers.append(outlier_data)

        print(f"Iteration {i+1}, Column {column_index}")
        print(f"Median: {np.median(amp_valid)}")
        print(f"Number of data points before filtering: {len(amp_valid)}")
        print(f"Number of data points after filtering: {num_chans_left}")
        print(f"Number of outliers: {outliers.sum()}")
        
        # 更新数据，去除异常值

    # 合并所有迭代中的异常值
    all_outliers = np.vstack(all_outliers)
    good_chans = np.setdiff1d(chans, all_outliers[:,0])
    # 保存两列数据，一列是频道号，一列是goodchans标记，1表示bad channel，0表示good channel
    allchans = np.where(np.isin(chans, good_chans), 0, 1)
    return allchans





def saveListFile(allList, fields, output):
    """
    save a list to a csv file
    """
    if len(fields) != len(allList):
        print("Size of allList and fields do not match!")
        sys.exit()

    allrows = []
    for i in range(allList[0].size):
        row = {}
        for j in range(len(fields)):
            row[fields[j]] = allList[j][i]
        allrows.append(row)

    csvfile = open(output, "w")
    obj = csv.DictWriter(csvfile, fieldnames=fields)
    obj.writeheader()
    obj.writerows(allrows)
    csvfile.close()


def readListFile(inFile, fields):
    """
    read a list from a csv file
    """
    csvfile = open(inFile, "r")
    obj = csv.DictReader(csvfile)
    fields_r = obj.fieldnames

    for term in fields:
        if term not in fields_r:
            print(term + 'not in fields!')
            sys.exit()
    
    allList = []
    for i in range(len(fields)):
        allList.append(np.array([0.]))

    num = 0
    for row in obj:
        for i in range(len(fields)):
            if num == 0:
                allList[i][num] = float(row[fields[i]])
            else:
                allList[i] = np.append(allList[i], float(row[fields[i]]))
        num += 1

    return allList


class Track:

    def __init__(self, posFile, kyFile=None, col_time='A', col_SDP_X='H', col_SDP_Y='I', col_SDP_Z='J'):
        """
        read in the excel file containing two sheets
        SDP_PhaPos_X, SDP_PhaPos_Y, SDP_PhasPos_Z will be used to calculate alt and az
        The corresponding columns are: H, I, and J.
        Time is in local time (Beijing time UTC+8) and the column is A. 
        
        Or read directly all the position information from posfile
        """
        self.output = posFile
        self.fields = ['MJD', 'RA2000', 'DEC2000', 'AZ', 'ALT']
        if kyFile is not None: 
            wb = load_workbook(kyFile)
            sheet = wb.worksheets[0]
            Nrows = sheet.max_row
            timeList = []
            mjd_sdp = np.zeros(Nrows)
            x_sdp = np.zeros(Nrows)
            y_sdp = np.zeros(Nrows)
            z_sdp = np.zeros(Nrows)
            nn = 0
            for row in range(2, Nrows+1):
                t = Time(sheet[col_time+str(row)].value.encode('ascii')) - UTC_OFFSET
                x = sheet[col_SDP_X+str(row)].value
                y = sheet[col_SDP_Y+str(row)].value
                z = sheet[col_SDP_Z+str(row)].value
                # exclude NaNs and 0 for x, y, z.
                if np.isnan(x) or np.isnan(y) or np.isnan(z) or x==0 or y==0 or z==0: continue
                mjd_sdp[nn] = t.mjd  
                x_sdp[nn] = x
                y_sdp[nn] = y
                z_sdp[nn] = z
                timeList.append(t)
                nn += 1

            self.t = timeList
            self.mjd = mjd_sdp[:nn]
            self.x = x_sdp[:nn]
            self.y = y_sdp[:nn]
            self.z = z_sdp[:nn]
            self.pressure = 0.
            self.temperature = 0.
        else:
            [self.mjd, self.ra, self.dec, self.az, self.alt] = readListFile(posFile, self.fields)
             

    def xyz2azalt(self):
        """
        convert x, y, z to az, alt
        Return: az and alt in degrees
        """
        R_ca = np.sqrt(self.x*self.x+self.y*self.y+self.z*self.z)
        az0 = np.degrees(np.arctan2(self.x, self.y)) % 360.
        alt = 90. - np.degrees(np.arccos(-self.z/R_ca))
        az = np.copy(az0)

        # +180 if in the range [0,180.]
        sel1 = az0 >= 0.
        sel2 = az0 <= 180.
        sel = sel1 * sel2
        az[sel] += 180.

        # -180 if in the range (180, 360]
        sel1 = az0 > 180.
        sel2 = az0 <= 360.
        sel = sel1 * sel2
        az[sel] -= 180.

        self.az = az
        self.alt = alt


    def azalt2eq(self):
        """
        Convert az, alt to ra, dec
        Return: ra and dec in degrees
        """
        frame = AltAz(obstime=self.t, location=FAST_loc, pressure=self.pressure, temperature=self.temperature)
        c = SkyCoord(alt = self.alt*u.deg, az = self.az*u.deg, frame = frame)
        c_new = c.transform_to('fk5')
        self.ra = c_new.ra.deg
        self.dec = c_new.dec.deg


    def plotRADec(self,output=None):
        fig = plt.figure()
        ax = fig.add_subplot(1,1,1)
        ax.plot(self.ra, self.dec) 
        ax.set_xlabel('RA')
        ax.set_ylabel('Dec')
        if output==None: plt.show()
        else: plt.savefig(output)


    def plotRADec_T(self, output=None):
        fig = plt.figure()
        ax1 = fig.add_subplot(2,1,1)
        #ax1.plot(self.mjd, self.ra, 'o') 
        ax1.plot(self.ra, 'o') 
        ax1.set_ylabel('RA')
        ax2 = fig.add_subplot(2,1,2)
        #ax2.plot(self.mjd, self.dec, 'o') 
        ax2.plot(self.dec, 'o') 
        ax2.set_ylabel('Dec')
        ax2.set_xlabel('MJD')
        if output==None: plt.show()
        else: plt.savefig(output)


    def runsave(self):
        """
        output as csv file
        """
        print(self.output)
        self.xyz2azalt()
        self.azalt2eq()
        allList = [self.mjd, self.ra, self.dec, self.az, self.alt]
        saveListFile(allList, self.fields, self.output)


    def runsaveList(self, outputList):
        """
        output as csv file
        """

        # outputList = [[filename, i0, i1],..]
        
        for term in outputList: 
            i0 = term[1]
            i1 = term[2]
            allList = [self.mjd[i0:i1], self.ra[i0:i1], self.dec[i0:i1], self.az[i0:i1], self.alt[i0:i1]]
            saveListFile(allList, self.fields, term[0])


def prepare_data(work_dir):
    file_list = natsorted(glob(f"{work_dir}/*.fits"))
    
    data_all, record_list, good_chans_list = [], [], []
    
    for file in file_list:
        good_chans = process_data(file, 1, kernel_size=10, iterations=5)
        good_chans_list.append(good_chans)
        
        data, record = find_obs_time(file)
        record_list.extend(record)
        data_all.append(data)
    record_list[0] = record_list[1] - (record_list[3] - record_list[2]) 

    
    # === 2. 数据转换为 NumPy 数组 ===
    good_chans_arr = np.array(good_chans_list)
    data_all_arr = np.array(data_all)
    
    # === 3. 绘制观测数据 ===
    fig, axes = plt.subplots(2, 1, figsize=(10, 6))
    ax = axes[0]
    im = ax.imshow(data_all_arr, aspect='auto')
    for i in range(0, len(record_list), 2):
        ax.scatter(record_list[i:i+2], [i // 2] * 2, color='red', s=1)
    
    ax.set_xlabel('Time')
    ax.set_ylabel('BEAM')
    ax.set_yticks(np.arange(data_all_arr.shape[0]))
    ax.set_yticklabels(np.arange(data_all_arr.shape[0]) + 1)

    record_list_new= (np.array(record_list)*2.5 + 3512).astype(int)

    # === 4. 保存并绘制坏信道图 ===
    np.savetxt(f"{work_dir}/good_chans.txt", good_chans_arr)
    
    ax = axes[1]
    ax.imshow(good_chans_arr, aspect='auto', cmap='gray', vmin=0, vmax=1,interpolation='none')
    ax.set_xlabel('Time')
    ax.set_ylabel('Beam')
    ax.set_yticks(np.arange(data_all_arr.shape[0]))
    ax.set_yticklabels(np.arange(data_all_arr.shape[0]) + 1)
    ax.set_title('1 means bad channel')
    plt.savefig(f"{work_dir}/show_all.png")
    plt.show()
    np.savetxt("record_list.txt", record_list)
    return record_list_new


def main():
    """Process one epoch observation and visualize the results."""
    
    # === 1. 数据加载 ===
    work_dir = "../data_ave-64/3C138_20241207"
    record_list = prepare_data(work_dir)

    # === 5. 处理天文轨迹数据 ===
    kyFile = "../data_orig/KY/3C138_2024_12_07_02_27_00_000.xlsx"
    filebase = os.path.basename(kyFile)
    posFile = os.path.join(work_dir, filebase.replace('_00_000.xlsx', '.csv'))
    outfile = os.path.join(work_dir, filebase.replace('_00_000.xlsx', '_scan_pts.dat'))
    
    beamList = np.arange(1, 19, dtype=int)  # 1~18
    
    # 检查并加载轨迹文件
    if not os.path.exists(posFile):
        track = Track(posFile, kyFile=kyFile)
        track.runsave()
    else:
        track = Track(posFile)
    
    # === 6. 计算不同波束的角度位置 ===
    source = "3C138"
    c0 = SkyCoord.from_name(source)
    RA_c, Dec_c = c0.ra.deg, c0.dec.deg
    RotAngle = Angle(0. * u.deg)
    fig, axes = plt.subplots(3,6, figsize=(6, 9))
    for beam in beamList:
        ss = f'M{beam:02d}'
        print(ss)
        
        dec_t = np.radians(track.dec)
        ra_t = np.radians(track.ra)
        
        if ss != 'M01':
            dec_t += BeamOffsetDict[ss][1] * BeamOffsetDict[ss][0].rad * np.sin((RotAngle + BeamOffsetDict[ss][3]).rad)
            ra_t += BeamOffsetDict[ss][2] * BeamOffsetDict[ss][0].rad * np.cos((RotAngle + BeamOffsetDict[ss][3]).rad) / np.cos(dec_t)
        
        dec, ra = np.degrees(dec_t), np.degrees(ra_t)
        c = SkyCoord(ra=ra*u.deg, dec=dec*u.deg)
        ang_sep = c0.separation(c).arcmin
        
        if ra.size != dec.size:
            print("Error: RA and Dec size mismatch!")

        # === 7. 绘制波束位置图 ===
        ax1 = axes[(beam-1)//6, (beam-1)%6]
        
        ax1.plot(ra, dec)
        ax1.set_xlim(RA_c - 0.5, RA_c + 0.5)
        ax1.set_ylim(Dec_c - 0.5, Dec_c + 0.5)
        # reverse x-axis
        ax1.invert_xaxis()
        ax1.set_title(f'{ss}')
    plt.savefig(f"{work_dir}/beam_positions.png")
    print(track.mjd)
        
    # === 8. 保存扫描点数据 ===
    with open(outfile, 'w') as f:
        for i, beam in enumerate(beamList):
            ss = f'M{beam:02d} {track.mjd[record_list[i * 2] - 1]}  {track.mjd[record_list[i * 2 + 1] - 1]}\n'
            f.write(ss)
    
    print(f"Scan points saved to {outfile}")
   
if __name__ == "__main__":
    main()



# %%
