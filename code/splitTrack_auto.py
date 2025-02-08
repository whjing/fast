#!/usr/bin/env python
# -*- coding: utf-8 -*-
################################################################################
#  This is the code to calibrate polarization data observed with the FAST      #
#  telescope                                                                   # 
#      The code:                                                               #
#         - measures calibration parameters from the input calibration scans   #
#         - calibrate ALL rpf files in the current directory                   #
#         - bin the data in df-MHz wide bands                                  #
#         - write sdfits output file.                                          #
#                                                                              #
#      It assumes the following scans are available:                           #
#         - an unpolarised flux calibrator (1934, 0407, and HydA only for now) #
#         - a polarised calibrator (0043-424 and 3C138 only for now).          #
################################################################################


from astropy.io import fits as pf
from astropy.modeling import models, fitting
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import numpy as np
import astropy.units as u
from astropy.time import Time
from astropy.coordinates import SkyCoord, ITRS, EarthLocation, AltAz
from astropy.stats import sigma_clip
from scipy import interpolate
import csv
import sys
import os
import glob
import re as RegEx
from astropy import wcs
from scipy.interpolate import griddata
#from pyslalib import slalib
from scipy import ndimage
import gc
from scipy.signal import find_peaks
import math


# FAST locations
FAST_lat = 25.+39./60.+10.6/3600.
FAST_lon = 106.+51./60.+24.0/3600.
FAST_alt = 1110.0288
FAST_loc = EarthLocation(lat=FAST_lat*u.deg, lon=FAST_lon*u.deg, height=FAST_alt*u.m) 
UTC_OFFSET = 8.*u.hour
VC = 299792458. 
DIAM = 300.

#from astroplan import Observer
#from pytz import timezone
#observer = Observer(name='FAST Telescope',
#                    location = FAST_loc,
#                    timezone = timezone('Asia/Shanghai'))

# In case, finals2000A cannot be downloaded
from astropy.utils import iers
#iers.conf.auto_download = False
#iers.IERS.iers_table = iers.IERS_A.open("/home/xiaohui/erhai/FAST/finals2000A.all")
#iers.IERS.iers_table = iers.IERS_A.open("/home/xhsun/Processing/FAST/finals2000A.all")

from astropy.coordinates import Angle

def saveListFile(allList, fields, output):
    """
    save a list to a csv file
    """
    if len(fields) != len(allList):
        print("Size of allList and fields do not match!")
        sys.exit()

    allrows = []
    for i in range(allList[0].size):
        row = {}
        for j in range(len(fields)):
            row[fields[j]] = allList[j][i]
        allrows.append(row)

    csvfile = open(output, "w")
    obj = csv.DictWriter(csvfile, fieldnames=fields)
    obj.writeheader()
    obj.writerows(allrows)
    csvfile.close()


def readListFile(inFile, fields):
    """
    read a list from a csv file
    """
    csvfile = open(inFile, "r")
    obj = csv.DictReader(csvfile)
    fields_r = obj.fieldnames

    for term in fields:
        if term not in fields_r:
            print(term + 'not in fields!')
            sys.exit()
    
    allList = []
    for i in range(len(fields)):
        allList.append(np.array([0.]))

    num = 0
    for row in obj:
        for i in range(len(fields)):
            if num == 0:
                allList[i][num] = float(row[fields[i]])
            else:
                allList[i] = np.append(allList[i], float(row[fields[i]]))
        num += 1

    return allList


def load_obsStart(fitsFile):
    with pf.open(fitsFile) as hdul:
        hdu1 = hdul[1]
        hdr = hdu1.header
        # print(repr(hdr))

        data = hdu1.data
    obsMJDStart = data['UTOBS'][0]
    return obsMJDStart

class Track:

    def __init__(self, posFile, fitsFile, kyFile=None, col_time='A', col_SDP_X='H', col_SDP_Y='I', col_SDP_Z='J'):
        """
        read in the excel file containing two sheets
        SDP_PhaPos_X, SDP_PhaPos_Y, SDP_PhasPos_Z will be used to calculate alt and az
        The corresponding columns are: H, I, and J.
        Time is in local time (Beijing time UTC+8) and the column is A. 
        
        Or read directly all the position information from posfile
        """
        self.output = posFile
        self.fitsFile = fitsFile
        self.fields = ['MJD', 'RA2000', 'DEC2000', 'AZ', 'ALT']
        if kyFile is not None: 
            wb = load_workbook(kyFile)
            sheet = wb.worksheets[0]
            Nrows = sheet.max_row
            timeList = []
            mjd_sdp = np.zeros(Nrows)
            x_sdp = np.zeros(Nrows)
            y_sdp = np.zeros(Nrows)
            z_sdp = np.zeros(Nrows)
            nn = 0
            for row in range(2, Nrows+1):
                t = Time(sheet[col_time+str(row)].value.encode('ascii')) - UTC_OFFSET
                x = sheet[col_SDP_X+str(row)].value
                y = sheet[col_SDP_Y+str(row)].value
                z = sheet[col_SDP_Z+str(row)].value
                # exclude NaNs and 0 for x, y, z.
                obsMJDStart = load_obsStart(fitsFile)
                if t.mjd <= obsMJDStart or np.isnan(x) or np.isnan(y) or np.isnan(z) or x==0 or y==0 or z==0: continue
                mjd_sdp[nn] = t.mjd  
                x_sdp[nn] = x
                y_sdp[nn] = y
                z_sdp[nn] = z
                timeList.append(t)
                nn += 1

            self.t = timeList
            self.mjd = mjd_sdp[:nn]
            self.x = x_sdp[:nn]
            self.y = y_sdp[:nn]
            self.z = z_sdp[:nn]
            self.pressure = 0.
            self.temperature = 0.
        else:
            [self.mjd, self.ra, self.dec, self.az, self.alt] = readListFile(posFile, self.fields)
             

    def xyz2azalt(self):
        """
        convert x, y, z to az, alt
        Return: az and alt in degrees
        """
        R_ca = np.sqrt(self.x*self.x+self.y*self.y+self.z*self.z)
        az0 = np.degrees(np.arctan2(self.x, self.y)) % 360.
        alt = 90. - np.degrees(np.arccos(-self.z/R_ca))
        az = np.copy(az0)

        # +180 if in the range [0,180.]
        sel1 = az0 >= 0.
        sel2 = az0 <= 180.
        sel = sel1 * sel2
        az[sel] += 180.

        # -180 if in the range (180, 360]
        sel1 = az0 > 180.
        sel2 = az0 <= 360.
        sel = sel1 * sel2
        az[sel] -= 180.

        self.az = az
        self.alt = alt


    def azalt2eq(self):
        """
        Convert az, alt to ra, dec
        Return: ra and dec in degrees
        """
        frame = AltAz(obstime=self.t, location=FAST_loc, pressure=self.pressure, temperature=self.temperature)
        c = SkyCoord(alt = self.alt*u.deg, az = self.az*u.deg, frame = frame)
        c_new = c.transform_to('fk5')
        self.ra = c_new.ra.deg
        self.dec = c_new.dec.deg


    def plotRADec(self,output=None):
        fig = plt.figure()
        ax = fig.add_subplot(1,1,1)
        ax.plot(self.ra, self.dec) 
        ax.set_xlabel('RA')
        ax.set_ylabel('Dec')
        if scanDir == 'RA':
            ax.hlines(self.cut1)
            ax.hlines(self.cut2)
        else:
            ax.vlines(self.cut1)
            ax.vlines(self.cut2)
        if output==None: plt.show()
        else: plt.savefig(output)


    def plotRADec_T(self, output=None):
        fig = plt.figure()
        ax1 = fig.add_subplot(2,1,1)
        #ax1.plot(self.mjd, self.ra, 'o') 
        ax1.plot(self.ra, 'o') 
        ax1.set_ylabel('RA')
        ax2 = fig.add_subplot(2,1,2)
        #ax2.plot(self.mjd, self.dec, 'o') 
        ax2.plot(self.dec, 'o') 
        ax2.set_ylabel('Dec')
        ax2.set_xlabel('MJD')
        if output==None: plt.show()
        else: plt.savefig(output)


    def runsave(self):
        """
        output as csv file
        """
        self.xyz2azalt()
        self.azalt2eq()
        allList = [self.mjd, self.ra, self.dec, self.az, self.alt]
        saveListFile(allList, self.fields, self.output)


    def runsaveList(self, outputList):
        """
        output as csv file
        """

        # outputList = [[filename, i0, i1],..]
        
        for term in outputList: 
            i0 = term[1]
            i1 = term[2]
            print('i0',i0)
            print('i1',i1)
            allList = [self.mjd[i0:i1], self.ra[i0:i1], self.dec[i0:i1], self.az[i0:i1], self.alt[i0:i1]]
            saveListFile(allList, self.fields, term[0])

    def outputsave(self, outbase,scanDir,Nscan,Ra_c,Dec_c,window_size):
        # 计数


        if scanDir == 'RA':
            data = self.ra
            RaDec_c = Ra_c
            outbase = outbase+'_R'
        else:
            data = self.dec
            RaDec_c =Dec_c
            outbase = outbase+'_D'

        if Nscan % 2 ==0 :
            max = np.max(data)
            min = np.min(data)
            print(f'The max of dec is {max}\n The min of dec is {min}')
            # 找到与20最接近的数值
            nearest = max if abs(max - RaDec_c) < abs(min - RaDec_c) else min


            # 计算最接近数值与20的差值
            r = abs(nearest - RaDec_c) - 0.1

            # 生成两组数值
            self.cut1 = RaDec_c + r
            self.cut2 = RaDec_c - r

            print(self.cut1,self.cut2)






def main():
    # kyDir = "/media/xiaohui/c71c04aa-d268-49a3-95e2-ff6b76108ae4/PT2021_0111/KY/"  
    kyDir = '/share/home/whjing/fast/KY/'
    kyFile = kyDir + "G182_3_2022_12_05_00_58_00_000.xlsx"
    fitsFile = '/share/home/whjing/fast/orig/G182_3/20221205/G182_3_MultiBeamOTF-M01_W_0001.fits'
    outbase = kyFile.replace('/KY/', '/file_redo-day-3/').replace('_00_000.xlsx', '_test')
    posFile = outbase + '.csv'
    
    scanDir = 'RA'#'DEC'
    Nscan = 5
    Ra_c = 84.187500
    Dec_c = 24.675000
    window_size = 9
    #outputList = [['Cygnus_Loop_R1.csv',1890,7911], ['Cygnus_Loop_R2.csv',8159,14173], ['Cygnus_Loop_R3.csv',14436,20457], ['Cygnus_Loop_R4.csv',20696,26705], ['Cygnus_Loop_R5.csv',26987,-1]]
    
    if not os.path.exists(posFile):
        track = Track(posFile, kyFile=kyFile, fitsFile = fitsFile)
        track.runsave()
        outputList = track.outputsave(outbase,scanDir,Nscan,Ra_c,Dec_c,window_size)
        print(outputList)
        
        if not os.path.exists(outputList[0][0]): track.runsaveList(outputList)
    else:
        track = Track(posFile, fitsFile = fitsFile)
        outputList = track.outputsave(outbase,scanDir,Nscan,Ra_c,Dec_c,window_size)
        if not os.path.exists(outputList[0][0]): track.runsaveList(outputList)
        
    track.plotRADec(output = f'{outbase}_radec.png')
    track.plotRADec_T(output = f'{outbase}_radec-T.png')

    for term in outputList:
        posFile = term[0]
        track = Track(posFile, fitsFile = fitsFile)
        track.plotRADec()
        track.plotRADec_T()


if __name__ == "__main__":
    main()

