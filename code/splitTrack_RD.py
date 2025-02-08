#!/usr/bin/env python
# -*- coding: utf-8 -*-
################################################################################
#  This is the code to calibrate polarization data observed with the FAST      #
#  telescope                                                                   # 
#      The code:                                                               #
#         - measures calibration parameters from the input calibration scans   #
#         - calibrate ALL rpf files in the current directory                   #
#         - bin the data in df-MHz wide bands                                  #
#         - write sdfits output file.                                          #
#                                                                              #
#      It assumes the following scans are available:                           #
#         - an unpolarised flux calibrator (1934, 0407, and HydA only for now) #
#         - a polarised calibrator (0043-424 and 3C138 only for now).          #
################################################################################


from astropy.io import fits as pf
from astropy.modeling import models, fitting
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import numpy as np
import astropy.units as u
from astropy.time import Time
from astropy.coordinates import SkyCoord, ITRS, EarthLocation, AltAz
from astropy.stats import sigma_clip
from scipy import interpolate
import csv
import sys
import os
import glob
import re as RegEx
from astropy import wcs
from scipy.interpolate import griddata
#from pyslalib import slalib
from scipy import ndimage
import gc
from scipy.signal import find_peaks
import math


# FAST locations
FAST_lat = 25.+39./60.+10.6/3600.
FAST_lon = 106.+51./60.+24.0/3600.
FAST_alt = 1110.0288
FAST_loc = EarthLocation(lat=FAST_lat*u.deg, lon=FAST_lon*u.deg, height=FAST_alt*u.m) 
UTC_OFFSET = 8.*u.hour
VC = 299792458. 
DIAM = 300.

#from astroplan import Observer
#from pytz import timezone
#observer = Observer(name='FAST Telescope',
#                    location = FAST_loc,
#                    timezone = timezone('Asia/Shanghai'))

# In case, finals2000A cannot be downloaded
from astropy.utils import iers
#iers.conf.auto_download = False
#iers.IERS.iers_table = iers.IERS_A.open("/home/xiaohui/erhai/FAST/finals2000A.all")
#iers.IERS.iers_table = iers.IERS_A.open("/home/xhsun/Processing/FAST/finals2000A.all")

from astropy.coordinates import Angle

def saveListFile(allList, fields, output):
    """
    save a list to a csv file
    """
    if len(fields) != len(allList):
        print("Size of allList and fields do not match!")
        sys.exit()

    allrows = []
    for i in range(allList[0].size):
        row = {}
        for j in range(len(fields)):
            row[fields[j]] = allList[j][i]
        allrows.append(row)

    csvfile = open(output, "w")
    obj = csv.DictWriter(csvfile, fieldnames=fields)
    obj.writeheader()
    obj.writerows(allrows)
    csvfile.close()


def readListFile(inFile, fields):
    """
    read a list from a csv file
    """
    csvfile = open(inFile, "r")
    obj = csv.DictReader(csvfile)
    fields_r = obj.fieldnames

    for term in fields:
        if term not in fields_r:
            print(term + 'not in fields!')
            sys.exit()
    
    allList = []
    for i in range(len(fields)):
        allList.append(np.array([0.]))

    num = 0
    for row in obj:
        for i in range(len(fields)):
            if num == 0:
                allList[i][num] = float(row[fields[i]])
            else:
                allList[i] = np.append(allList[i], float(row[fields[i]]))
        num += 1

    return allList


class Track:

    def __init__(self, posFile, kyFile=None, col_time='A', col_SDP_X='H', col_SDP_Y='I', col_SDP_Z='J'):
        """
        read in the excel file containing two sheets
        SDP_PhaPos_X, SDP_PhaPos_Y, SDP_PhasPos_Z will be used to calculate alt and az
        The corresponding columns are: H, I, and J.
        Time is in local time (Beijing time UTC+8) and the column is A. 
        
        Or read directly all the position information from posfile
        """
        self.output = posFile
        self.fields = ['MJD', 'RA2000', 'DEC2000', 'AZ', 'ALT']
        if kyFile is not None: 
            wb = load_workbook(kyFile)
            sheet = wb.worksheets[0]
            Nrows = sheet.max_row
            timeList = []
            mjd_sdp = np.zeros(Nrows)
            x_sdp = np.zeros(Nrows)
            y_sdp = np.zeros(Nrows)
            z_sdp = np.zeros(Nrows)
            nn = 0
            for row in range(2, Nrows+1):
                t = Time(sheet[col_time+str(row)].value.encode('ascii')) - UTC_OFFSET
                x = sheet[col_SDP_X+str(row)].value
                y = sheet[col_SDP_Y+str(row)].value
                z = sheet[col_SDP_Z+str(row)].value
                # exclude NaNs and 0 for x, y, z.
                if np.isnan(x) or np.isnan(y) or np.isnan(z) or x==0 or y==0 or z==0: continue
                mjd_sdp[nn] = t.mjd  
                x_sdp[nn] = x
                y_sdp[nn] = y
                z_sdp[nn] = z
                timeList.append(t)
                nn += 1

            self.t = timeList
            self.mjd = mjd_sdp[:nn]
            self.x = x_sdp[:nn]
            self.y = y_sdp[:nn]
            self.z = z_sdp[:nn]
            self.pressure = 0.
            self.temperature = 0.
        else:
            [self.mjd, self.ra, self.dec, self.az, self.alt] = readListFile(posFile, self.fields)
             

    def xyz2azalt(self):
        """
        convert x, y, z to az, alt
        Return: az and alt in degrees
        """
        R_ca = np.sqrt(self.x*self.x+self.y*self.y+self.z*self.z)
        az0 = np.degrees(np.arctan2(self.x, self.y)) % 360.
        alt = 90. - np.degrees(np.arccos(-self.z/R_ca))
        az = np.copy(az0)

        # +180 if in the range [0,180.]
        sel1 = az0 >= 0.
        sel2 = az0 <= 180.
        sel = sel1 * sel2
        az[sel] += 180.

        # -180 if in the range (180, 360]
        sel1 = az0 > 180.
        sel2 = az0 <= 360.
        sel = sel1 * sel2
        az[sel] -= 180.

        self.az = az
        self.alt = alt


    def azalt2eq(self):
        """
        Convert az, alt to ra, dec
        Return: ra and dec in degrees
        """
        frame = AltAz(obstime=self.t, location=FAST_loc, pressure=self.pressure, temperature=self.temperature)
        c = SkyCoord(alt = self.alt*u.deg, az = self.az*u.deg, frame = frame)
        c_new = c.transform_to('fk5')
        self.ra = c_new.ra.deg
        self.dec = c_new.dec.deg


    def plotRADec(self,output=None):
        fig = plt.figure()
        ax = fig.add_subplot(1,1,1)
        ax.plot(self.ra, self.dec) 
        ax.set_xlabel('RA')
        ax.set_ylabel('Dec')
        if output==None: plt.show()
        else: plt.savefig(output)


    def plotRADec_T(self, output=None):
        fig = plt.figure()
        ax1 = fig.add_subplot(2,1,1)
        #ax1.plot(self.mjd, self.ra, 'o') 
        ax1.plot(self.ra, 'o') 
        ax1.set_ylabel('RA')
        ax2 = fig.add_subplot(2,1,2)
        #ax2.plot(self.mjd, self.dec, 'o') 
        ax2.plot(self.dec, 'o') 
        ax2.set_ylabel('Dec')
        ax2.set_xlabel('MJD')
        if output==None: plt.show()
        else: plt.savefig(output)


    def runsave(self):
        """
        output as csv file
        """
        self.xyz2azalt()
        self.azalt2eq()
        allList = [self.mjd, self.ra, self.dec, self.az, self.alt]
        saveListFile(allList, self.fields, self.output)


    def runsaveList(self, outputList):
        """
        output as csv file
        """

        # outputList = [[filename, i0, i1],..]
        
        for term in outputList: 
            i0 = term[1]
            i1 = term[2]
            print('i0',i0)
            print('i1',i1)
            allList = [self.mjd[i0:i1], self.ra[i0:i1], self.dec[i0:i1], self.az[i0:i1], self.alt[i0:i1]]
            saveListFile(allList, self.fields, term[0])

    def outputsave(self, infTemplate,outbase,scanDir,Nscan,Ra_c,Dec_c,window_size):
        def Value(Tt,window_size):
            size = window_size
            min_range = int(Tt * 0.0005)
            max_range = int(Tt * 0.002)
            print('min_range:',min_range)
            print('max_range:',max_range)

            # 找到最小和最大的奇数
            if min_range % 2 == 0:
                min_range = min_range+1
            min_range = max(window_size,min_range)
            print('max(window_size,min_range):',min_range)

            def findvalue(min_range,max_range,window_size,size):
                for i in range(min_range,max_range,2):
                    if Tt % i == 0 and min_range > size:
                        window_size = i
                    else:
                        continue
                return window_size

            for j in range(10):
                print('Tt:',Tt)
                print('window_size:',window_size)
                window_size = findvalue(min_range,max_range,window_size,size)
                print('new_window_size:',window_size)
                if window_size == size:
                    sv = size +2
                    Tt = math.floor(Tts / sv) * sv#size+2
                    print('newTt:',Tt)
                    return sv, Tt
                else:
                    return window_size,Tt

        #预设参数
        #Nscan = 5
        Npeak = 3*Nscan-2
        #Ra_c = 296.27
        #window_size = -1#起步值？？？

        # 加载数据
        #data = np.loadtxt("C:/Users/23627/Desktop/RA.dat")
        if scanDir == 'RA':
            data = self.ra
            RaDec_c = Ra_c
            outbase = outbase+'_R'
        else:
            data = self.dec
            RaDec_c =Dec_c
            outbase = outbase+'_D'

        da = len(data)
        #
        kyTime = self.mjd
        infdata = pf.getdata(infTemplate)
        ObsTime = infdata['UTOBS']
        Tmin = np.min(ObsTime)
        y_index = next((i for i, v in enumerate(kyTime) if v > Tmin),None)
        print('y_idex:',y_index)
        y0 = data[y_index]
        y1 = data[y_index+2*window_size]#两倍步长内存在第一个峰值与否
        
        # if y0 > y1:
        #     peak0 = False
        # else:
        #     peak0 = True

        if y0>y1 and y0>RaDec_c:
            peak0=False
            dir0=1
        elif y0>y1 and y0<RaDec_c:
            peak0=True
            dir0=0
        if y0<y1 and y0<RaDec_c:
            peak0 = False
            dir0 = 0
        if y0<y1 and y0>RaDec_c:
            peak0=True
            dir0=1





        #sys.exit()
        #y_index = np.argmax(data > RaDec_c)

        # 删除第一次超过阈值之前的所有数据
        y = data[y_index:]
        print(da)
        x = np.arange(len(y))
        Tts = len(x)

        # 每10个点为步长求一次平均
        def Peak(x,y,window_size):
            smoothed_x = np.mean(x.reshape(-1, window_size), axis=1)
            smoothed_y = np.mean(y.reshape(-1, window_size), axis=1)

            # 寻找极大值和极小值
            peaks, _ = find_peaks(smoothed_y)
            valleys, _ = find_peaks(-smoothed_y)
        #     print(len(peaks))
        #     print(len(valleys))
            print(len(peaks)+len(valleys))
            return peaks,valleys,smoothed_x,smoothed_y

        window_size,Tt= Value(Tts,window_size)
        print('1111',Tt)
        y = data[da-Tt:]
        x = np.arange(len(y))
        peaks,valleys,smoothed_x,smoothed_y= Peak(x,y,window_size)
        for j in range(10):
            if peak0 ==False:
                if len(peaks)+len(valleys)>Npeak+1:
                    window_size,Tt= Value(Tt,window_size)
                    y = data[da-Tt:]
                    x = np.arange(len(y))
                    peaks,valleys,smoothed_x,smoothed_y = Peak(x,y,window_size)
                    print('new',Tt)
                    continue
                else:break
            else:
                if len(peaks)+len(valleys)>Npeak:
                    window_size,Tt= Value(Tt,window_size)
                    y = data[da-Tt:]
                    x = np.arange(len(y))
                    peaks,valleys,smoothed_x,smoothed_y = Peak(x,y,window_size)
                    print('new',Tt)
                    continue
                else:break
        print("奇数因数:", window_size)

        # 获取极值点及其位置
        print(Tts)
        print(Tt)
        top_peaks = [(smoothed_x[peak], smoothed_y[peak]) for peak in peaks]
        bottom_valleys = [(smoothed_x[valley], smoothed_y[valley]) for valley in valleys]

        # 绘制散点图和连接曲线
        plt.figure(figsize=(10, 6))
        plt.scatter(x, y, color='dodgerblue', alpha=0.5, label='Original Data')
        plt.plot(smoothed_x, smoothed_y, color='tomato', label='Smoothed Curve')
        plt.plot(smoothed_x[peaks], smoothed_y[peaks], 'rx', label='Top Peaks')
        plt.plot(smoothed_x[valleys], smoothed_y[valleys], 'bo', label='Bottom Valleys')

        # 在图上标记极值点
        for position, value in top_peaks:
            plt.annotate(f'({position:.2f}, {value:.2f})', xy=(position, value), xytext=(position + 1, value + 1),
                        arrowprops=dict(facecolor='red', shrink=0.05), fontsize=8, color='red')

        for position, value in bottom_valleys:
            plt.annotate(f'({position:.2f}, {value:.2f})', xy=(position, value), xytext=(position + 1, value - 1),
                        arrowprops=dict(facecolor='blue', shrink=0.05), fontsize=8, color='blue')

        plt.xlabel('Time')  # 横轴表示时间
        plt.ylabel('Ra/Dec')  # 纵轴表示 Ra/Dec
        plt.title('Scatter Plot with Averaged Data and Connecting Curve')
        plt.legend()
        plt.grid(True)
        plt.savefig(f'{outbase}_splitTrack.png')
        #plt.show()

        print(top_peaks)
        print(bottom_valleys)
        print("极大值及其位置：")
        for i, (position, value) in enumerate(top_peaks, 1):
            print(f"第{i}个极大值位置：{position+da-Tt:.2f}，对应值：{value:.2f}")

        print("\n极小值及其位置：")
        for i, (position, value) in enumerate(bottom_valleys, 1):
            print(f"第{i}个极小值位置：{position+da-Tt:.2f}，对应值：{value:.2f}")



        if dir0 == 1:
            v=top_peaks
            p=bottom_valleys
        else:
            p=top_peaks
            v=bottom_valleys

        p_values = [pair[0] for pair in p]
        p_Ra_values = [pair[1] for pair in p]

        v_values = [pair[0] for pair in v]
        v_Ra_values = [pair[1] for pair in v]

        # 合并 p 和 v 中的第一个数按 p1，v1，p2，v2，p3，v3 的顺序交替合成一个新的列表
        merged_values = []
        for i in range(min(len(p), len(v))):
            merged_values.append(p_values[i])
            merged_values.append(v_values[i])

        # 如果 p 或 v 中有剩余的元素，则添加到 merged_values 中
        if len(p) > len(v):
            merged_values.extend(p_values[len(v):])
        elif len(v) > len(p):
            merged_values.extend(v_values[len(p):])

        # 合并 p 和 v 中的第二个数按 p1，v1，p2，v2，p3，v3 的顺序交替合成一个新的列表
        merged_Ra_values = []
        for i in range(min(len(p), len(v))):
            merged_Ra_values.append(p_Ra_values[i])
            merged_Ra_values.append(v_Ra_values[i])

        # 如果 p 或 v 中有剩余的元素，则添加到 merged_Ra_values 中
        if len(p) > len(v):
            merged_Ra_values.extend(p_Ra_values[len(v):])
        elif len(v) > len(p):
            merged_Ra_values.extend(v_Ra_values[len(p):])

        print(merged_values)
        print(merged_Ra_values)

        # 遍历 merged_values，判断 Ra_c/Dec_c 是否在相邻的第一个数和第二个数之间，如果是则保存到 out 列表中
        out = []
        Rn = 0
        for i in range(0, len(merged_values)):
            if peak0 == False:
                if i == 0:
                    Rn = Rn + 1
                    out.append([outbase + str(Rn)+'.csv', int(y_index), int(merged_values[0]+int(da-Tt))])
                elif i+1 < len(merged_values):
                    if merged_Ra_values[i] < RaDec_c < merged_Ra_values[i + 1] or merged_Ra_values[i] > RaDec_c > merged_Ra_values[i + 1]:
                        Rn = Rn + 1
                        out.append([outbase + str(Rn)+'.csv', int(merged_values[i]+int(da-Tt)), int(merged_values[i+1]+int(da-Tt))])
                else:
                    Rn = Rn + 1
                    out.append([outbase + str(Rn)+'.csv', int(merged_values[i]+int(da-Tt)), -1])
            else:
                if i+1 < len(merged_values):
                    if merged_Ra_values[i] < RaDec_c < merged_Ra_values[i + 1] or merged_Ra_values[i] > RaDec_c > merged_Ra_values[i + 1]:
                        Rn = Rn + 1
                        out.append([outbase + str(Rn)+'.csv', int(merged_values[i]+int(da-Tt)), int(merged_values[i+1]+int(da-Tt))])
                        #out.append(['R' + str(i // 2 + 2), merged_values[i + 1], merged_values[i + 2]])
                    #elif RaDec_c < merged_Ra_values[i]:
                        #out.append(['R' + str(i // 2 + 1), merged_values[i], merged_values[i+1]])
                else:
                    Rn = Rn + 1
                    out.append([outbase + str(Rn)+'.csv', int(merged_values[i]+int(da-Tt)), -1])

        print(out)
        return out




def main():
    # kyDir = "/media/xiaohui/c71c04aa-d268-49a3-95e2-ff6b76108ae4/PT2021_0111/KY/"  

    infTemplate = "../data_ave-64/G159+7_4_20241207/G159+7_4_MultiBeamOTF-M01_ave-64.fits"

    kyFile = "../data_ave-64/KY/G159+7_4_2024_12_07_00_08_00_000.xlsx"

    outbase = "../data_ave-64/G159+7_4_20241207/G159+7_4_2024_12_07_00_08"
    posFile = "../data_ave-64/G159+7_4_20241207/G159+7_4_2024_12_07_00_08.csv"
    
    scanDir = 'RA'#'DEC' --
    Nscan = 5
    Ra_c = 80.
    Dec_c = 50.
    window_size = 9
    #outputList = [['Cygnus_Loop_R1.csv',1890,7911], ['Cygnus_Loop_R2.csv',8159,14173], ['Cygnus_Loop_R3.csv',14436,20457], ['Cygnus_Loop_R4.csv',20696,26705], ['Cygnus_Loop_R5.csv',26987,-1]]
    
    if not os.path.exists(posFile):
        track = Track(posFile, kyFile=kyFile)
        track.runsave()
        outputList = track.outputsave(infTemplate,outbase,scanDir,Nscan,Ra_c,Dec_c,window_size)
        print(outputList)
        if not os.path.exists(outputList[0][0]): track.runsaveList(outputList)
    else:
        track = Track(posFile)
        outputList = track.outputsave(infTemplate,outbase,scanDir,Nscan,Ra_c,Dec_c,window_size)
        if not os.path.exists(outputList[0][0]): track.runsaveList(outputList)
    track.plotRADec()
    track.plotRADec_T()

    for term in outputList:
        posFile = term[0]
        track = Track(posFile)
        track.plotRADec()
        track.plotRADec_T()


if __name__ == "__main__":
    main()

