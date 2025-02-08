#!/usr/bin/env python
################################################################################
#  This is the code to calibrate polarization data observed with the FAST      #
#  telescope                                                                   # 
#      The code:                                                               #
#         - measures calibration parameters from the input calibration scans   #
#         - calibrate ALL rpf files in the current directory                   #
#         - bin the data in df-MHz wide bands                                  #
#         - write sdfits output file.                                          #
#                                                                              #
#      It assumes the following scans are available:                           #
#         - an unpolarised flux calibrator (1934, 0407, and HydA only for now) #
#         - a polarised calibrator (0043-424 and 3C138 only for now).          #
################################################################################


from astropy.io import fits as pf
from astropy.modeling import models, fitting
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import numpy as np
import astropy.units as u
from astropy.time import Time
from astropy.coordinates import SkyCoord, ITRS, EarthLocation, AltAz
from astropy.stats import sigma_clip
from scipy import interpolate
import csv
import sys
import os
import glob
import re as RegEx
#from kapteyn import maputils
from astropy import wcs
from scipy.interpolate import griddata
#from pyslalib import slalib
from scipy import ndimage
import gc
import pyregion


# FAST locations
FAST_lat = 25.+39./60.+10.6/3600.
FAST_lon = 106.+51./60.+24.0/3600.
FAST_alt = 1110.0288
FAST_loc = EarthLocation(lat=FAST_lat*u.deg, lon=FAST_lon*u.deg, height=FAST_alt*u.m) 
UTC_OFFSET = 8.*u.hour
VC = 299792458. 
DIAM = 300.

#from astroplan import Observer
#from pytz import timezone
#observer = Observer(name='FAST Telescope',
#                    location = FAST_loc,
#                    timezone = timezone('Asia/Shanghai'))

# In case, finals2000A cannot be downloaded
from astropy.utils import iers
#iers.conf.auto_download = False
#iers.IERS.iers_table = iers.IERS_A.open("/home/xiaohui/erhai/FAST/finals2000A.all")
#iers.IERS.iers_table = iers.IERS_A.open("/home/xhsun/Processing/FAST/finals2000A.all")

from astropy.coordinates import Angle
BeamOffsetDict={}
BeamOffsetDict['M08']= [Angle(11.69*u.arcmin), 1.0, 1.0, Angle(0.*u.deg)] 
BeamOffsetDict['M02']= [Angle(5.89*u.arcmin), 1.0, 1.0, Angle(0.*u.deg)] 
BeamOffsetDict['M05']= [Angle(5.93*u.arcmin), -1.0, -1.0, Angle(0.*u.deg)] 
BeamOffsetDict['M14']= [Angle(11.78*u.arcmin), -1.0, -1.0, Angle(0.*u.deg)] 

BeamOffsetDict['M19']= [Angle(10.14*u.arcmin), 1.0, 1.0, Angle(30.*u.deg)] 
BeamOffsetDict['M13']= [Angle(10.17*u.arcmin), -1.0, -1.0, Angle(30.*u.deg)] 

BeamOffsetDict['M09']= [Angle(10.2*u.arcmin), 1.0, 1.0, Angle(-30.*u.deg)] 
BeamOffsetDict['M15']= [Angle(10.07*u.arcmin), -1.0, -1.0, Angle(-30.*u.deg)] 

BeamOffsetDict['M10']= [Angle(11.74*u.arcmin), 1.0, 1.0, Angle(-60.*u.deg)] 
BeamOffsetDict['M03']= [Angle(5.885*u.arcmin), 1.0, 1.0, Angle(-60.*u.deg)] 
BeamOffsetDict['M06']= [Angle(5.835*u.arcmin), -1.0, -1.0, Angle(-60.*u.deg)] 
BeamOffsetDict['M16']= [Angle(11.67*u.arcmin), -1.0, -1.0, Angle(-60.*u.deg)] 

BeamOffsetDict['M18']= [Angle(11.775*u.arcmin), 1.0, 1.0, Angle(60.*u.deg)] 
BeamOffsetDict['M07']= [Angle(5.92*u.arcmin), 1.0, 1.0, Angle(60.*u.deg)] 
BeamOffsetDict['M04']= [Angle(5.76*u.arcmin), -1.0, -1.0, Angle(60.*u.deg)] 
BeamOffsetDict['M12']= [Angle(11.645*u.arcmin), -1.0, -1.0, Angle(60.*u.deg)] 

BeamOffsetDict['M17']= [Angle(10.15*u.arcmin), -1.0, -1.0, Angle(-90.*u.deg)] 
BeamOffsetDict['M11']= [Angle(10.15*u.arcmin), 1.0, 1.0, Angle(-90.*u.deg)] 


def ArPLS(y, lam, ratio, itermax):
    N = len(y)
    D = sparse.eye(N, format='csc')
    D = D[1:] - D[:-1]
    D = D[1:] - D[:-1]
    D = D.T
    w = np.ones(N)
    for i in range(itermax):
        W = sparse.diags(w, 0, shape=(N, N))
        Z = W + lam * D.dot(D.T)
        z = spsolve(Z, w * y)
        d = y - z
        dn = d[d < 0]
        m = np.mean(dn)
        s = np.std(dn)
        wt = 1. / (1 + np.exp(2 * (d - (2 * s - m)) / s))
        if np.linalg.norm(w - wt) / np.linalg.norm(w) < ratio:
            break
        w = wt
    return z


def getMAD(data):
    med = np.ma.median(data)
    mad = np.ma.median(abs(data - med))
    return mad


def RFI_flag(infTemplate, fileNr, BeamList, ChanFlag = [[0, 65535]], filt_size = 1000, NChan = 65536, Nsig=5, check=True):

    for beam in BeamList:
        fout = open("rms_b%02d.dat"%beam, "w")
        fout2 = open("good_chan_b%02d.dat"%beam, "w")

        print("Beam: ",beam)
        BeamNr = "M%02d" % beam 
        inf = infTemplate.replace('YYY', BeamNr[1:])

        rms = [[], [], [], []]
        rmsC = []

        for Chan in range(NChan):
            if Chan % 1000 == 0: print(Chan)
            for i in range(len(fileNr)):
                data = pf.getdata(inf.replace('XXX',fileNr[i]))
                if i==0: 
                    ObsData = data['DATA'][:,Chan,:]/1.e8
                else:
                    ObsData = np.concatenate((ObsData, data['DATA'][:,Chan,:]/1.e8),axis=0)
            ss = ""
            for j in range(len(rms)):
                rms_t = np.std(ObsData[:,j])
                rms[j].append(rms_t)
                ss += str(rms_t)+" "
            fout.write(str(Chan) + " " + ss+"\n")
            rmsC.append(Chan)
        fout.close()

        rmsC = np.array(rmsC)
        good_flag = np.zeros(rmsC.size)
        good_flag[ChanFlag[0][0]:ChanFlag[0][1]+1] = 1
        for i in range(len(ChanFlag)-1): good_flag[ChanFlag[i+1][0]:ChanFlag[i+1][1]+1] = 0

        medfit = []
        for i in range(len(rms)):
            rms[i] = np.array(rms[i])
            tmp = rsm[i] * 1.
            result = ndimage.median_filter(tmp, size=filt_size)
            diff = tmp - result
            mad = getMAD(diff[good_flag > 0.5])
            sel_t = abs(diff) > Nsig * mad
            good_flag[sel_t] = 0
            medfit.append(result)

        for i in range(good_flag.size):
            fout2.write(str(i)+" "+str(good_flag[i])+"\n")
        fout2.close()

        if check:
            fig = plt.figure()
            for i in range(len(rms)):
                ax1 = fig.add_subplot(2,2,i+1)
                ax1.plot(rmsC, rms[i], 'o')
                ax1.plot(rmsC[good_flag>0.5], rms[i][good_flag>0.5], 'o')
                ax1.plot(rmsC, medfit[i])
            plt.show()


def combineBeams(hdr, output, stokes='I'):
    """
    combine all the beams
    """
    data = np.zeros((hdr['NAXIS2'], hdr['NAXIS1']))
    w0 = wcs.WCS(hdr)
    
    for i in range(hdr['NAXIS2']):
        for j in range(hdr['NAXIS1']):
            world = w0.wcs_pix2world([[j,i]],0) 
            for k in range(1, 20): 
                inf = 'M%02d_%s.fits' % (k, stokes)
                hdr1 = pf.getheader(inf)
                data1 = pf.getdata(inf)
                w1 = wcs.WCS(hdr1)
                pix = w1.wcs_world2pix(world,0)
                if (0 <= int(pix[0][0]) <= hdr1['NAXIS1']-1) and (0 <= int(pix[0][1]) <= hdr1['NAXIS2']-1):
                    if not np.isnan(data1[int(pix[0][1]), int(pix[0][0])]):
                        data[i,j] += data1[int(pix[0][1]), int(pix[0][0])]

    pf.writeto(output, data, hdr, overwrite=True)


def plotImage(ax, inf, glvisible=True, gbvisible=False, gltickvisible=True, gbtickvisible=True):
    f_data = maputils.FITSimage(inf)
    tmp_nNaN = ~np.isnan(f_data.dat)
    tmp_data = f_data.dat[tmp_nNaN]
    clipmin=np.percentile(tmp_data.flatten(),1.)
    clipmax=np.percentile(tmp_data.flatten(),99.)
    annim = f_data.Annotatedimage(ax, clipmin=clipmin, clipmax=clipmax)
    annim.Image(interpolation='spline36')
    grat = annim.Graticule()
    grat.setp_axislabel("bottom", visible=glvisible)
    grat.setp_axislabel("left", visible=gbvisible)
    grat.setp_ticklabel(wcsaxis=0, visible=gltickvisible)
    grat.setp_ticklabel(wcsaxis=1, visible=gbtickvisible)
    annim.plot()

def saveListFile(allList, fields, output):
    """
    save a list to a csv file
    """
    if len(fields) != len(allList):
        print("Size of allList and fields do not match!")
        sys.exit()

    allrows = []
    for i in range(allList[0].size):
        row = {}
        for j in range(len(fields)):
            row[fields[j]] = allList[j][i]
        allrows.append(row)

    csvfile = open(output, "w")
    obj = csv.DictWriter(csvfile, fieldnames=fields)
    obj.writeheader()
    obj.writerows(allrows)
    csvfile.close()


def readListFile(inFile, fields):
    """
    read a list from a csv file
    """
    csvfile = open(inFile, "r")
    obj = csv.DictReader(csvfile)
    fields_r = obj.fieldnames

    for term in fields:
        if term not in fields_r:
            print(term + 'not in fields!')
            sys.exit()
    
    allList = []
    for i in range(len(fields)):
        allList.append(np.array([0.]))

    num = 0
    for row in obj:
        for i in range(len(fields)):
            if num == 0:
                allList[i][num] = float(row[fields[i]])
            else:
                allList[i] = np.append(allList[i], float(row[fields[i]]))
        num += 1

    return allList


def flux_model(source, freq):
    """
    get flux density (Jy) of a calibrator at a given frequency (GHz) 
    The model is from Perley & Butler (2017)
    lg(S) = a0 + a1*lg(f) + a2*(lgf)**2 + a3*(lgf)**3 ...
    """
    sou = {}
    sou['3C48'] = [1.3253, -0.7553, -0.1914, 0.0498]
    sou['3C138'] = [1.0088, -0.4981, -0.1552, -0.0102, 0.0223]
    sou['3C286'] = [1.2481, -0.4507, -0.1798, 0.0357]

    if source.upper() not in sou:
        print("Calibrator ", source.upper(), " not availble!")
        return None 
    else:
        value = sou[source.upper()]
        s = 0.
        lgf = np.log10(freq)
        for i in range(len(value)):
            s += value[i] * lgf ** float(i) 
        return 10**s


def pol_model(source,freq):
    """
    get polarization percentage and polarization angle
    Based on Perley & Butler (2013)
    freq in GHz
    """
    if source.upper() == '3C286':
        #pc = 8.6 # 1.05 GHz
        #pc = 9.5 # 1.45 GHz
        pc = 9.1 # assume an average
        pa = 33.
    elif source.upper() == '3C138':
        #pc = 5.6 # 1.05 GHz
        #pc = 7.5 # 1.45 GHz
        pc = 6.6 # assume an average
        #pa = -11  # 1.45 GHz
        #pa = -14  # 1.05 GHz
        pa = -12.5 # assume an average
    elif source.upper() == '3C48':
        #pc = 0.3 # 1.05 GHz
        #pc = 0.5 # 1.45 GHz
        pc = 0.4 # assume an average
        #pa = 25 # 1.05 GHz
        #pa = 140 # 1.45 GHz
        pa = 25 # very uncertain

    return pc, pa



def expsinc(x, y, x0, diam, fghz):
    """
    Make interpolation weighted by exp*sinc
    Mangum, Emerson, Greisen (2007) A&A 474, 679
    sinc(pi * z / a) * exp ( -|z|^2 / b^2 )
    a: 1.55 * beam/3
    b: 2.52 * beam/3
    """
    lam = VC/(fghz*1.e9)  # in m
    beam = lam / diam

    a = 1.55 * beam / 3. 
    b = 2.52 * beam / 3. 

    y0 = np.zeros_like(x0)
    for i in range(x0.size):
        dx = np.radians(x - x0[i])
        weights = np.sinc(np.pi*dx/a) * np.exp(-(dx*dx)/(b*b)) 
        y0[i] = np.sum(weights * y) / np.sum(weights)

    return y0
    

def parallactic_angle(mjd, target, loc):
    """
    Calculate the parallactic angle.
    Parameters
    ----------
    time : `~astropy.time.Time`
        Observation time.
    target : `~astroplan.FixedTarget` or `~astropy.coordinates.SkyCoord` or list
        Target celestial object(s).
    Returns
     -------
        parallactic angle

    Notes
    -----
    The parallactic angle is the angle between the great circle that
    intersects a celestial object and the zenith, and the object's hour
    circle [1]_.
    .. [1] https://en.wikipedia.org/wiki/Parallactic_angle
    """
    # Eqn (14.1) of Meeus' Astronomical Algorithms
    time = Time(mjd, format='mjd')

    """
    LST = time.sidereal_time('mean', longitude=loc.lon)
    H = (LST - target.ra).radian
    q = np.arctan2(np.sin(H),
                   (np.tan(loc.lat.radian) *
                   np.cos(target.dec.radian) -
                   np.sin(target.dec.radian)*np.cos(H)))*u.rad

    """
    c_ITRS = target.transform_to(ITRS(obstime=time))
    # Calculate local apparent Hour Angle (HA), wrap at 0/24h
    local_ha = loc.lon - c_ITRS.spherical.lon
    local_ha.wrap_at(24*u.hourangle, inplace=True)
    H = local_ha.radian
    # Calculate local apparent Declination
    local_dec = c_ITRS.spherical.lat

    q = np.arctan2(np.sin(H),
                   (np.tan(loc.lat.radian) *
                   np.cos(local_dec.radian) -
                   np.sin(local_dec.radian)*np.cos(H)))*u.rad

    return q.value


def parallactic_angle_sla(mjd, target, loc):
    """
    Calcualte the parallactic angle with slalib
    Have checked that the results are consistent with that from above
    """
    para = np.zeros_like(mjd)
    epoch = 2000.0

    for i in range(mjd.size):
        # auxiliary array
        amprms = slalib.sla_mappa(epoch, mjd[i])
        # calculate ra & dec to date
        ra_0 = target.ra.rad[i]
        dec_0 = target.dec.rad[i]
        ra_a, dec_a = slalib.sla_mapqkz(ra_0, dec_0, amprms)
        aob, zob, hob, dob, rob = slalib.sla_aop(ra_a, dec_a, mjd[i], 0.0, loc.lon.radian, loc.lat.radian, loc.height.value, 0.0, 0.0, 273.1550, 0.0, 0.0, 36000.0, 0.0065)
        para[i] = slalib.sla_pa(hob, dec_a, loc.lat.radian) 

    return para


def get_mask(file_region, data):
    print(f'the shape of orignal data is {data.shape}')
    reg = pyregion.open(file_region)
    data_mask = ~reg.get_mask(shape=data.shape)
    data_tmp = np.where(data_mask, np.nan, np.squeeze(data))
    print(f'the shape of template data is {data_tmp.shape}')
    # plt.imshow(data_tmp)
    # plt.show()
    # plt.close()
    # # pf.writeto(file_tmp, data = data, header=hdr)
    return data_tmp

    

class Track:

    def __init__(self, posFile, kyFile=None, col_time='A', col_SDP_X='H', col_SDP_Y='I', col_SDP_Z='J'):
        """
        read in the excel file containing two sheets
        SDP_PhaPos_X, SDP_PhaPos_Y, SDP_PhasPos_Z will be used to calculate alt and az
        The corresponding columns are: H, I, and J.
        Time is in local time (Beijing time UTC+8) and the column is A. 
        
        Or read directly all the position information from posfile
        """
        self.output = posFile
        self.fields = ['MJD', 'RA2000', 'DEC2000', 'AZ', 'ALT']
        if kyFile is not None: 
            wb = load_workbook(kyFile)
            sheet = wb.worksheets[0]
            Nrows = sheet.max_row
            timeList = []
            mjd_sdp = np.zeros(Nrows)
            x_sdp = np.zeros(Nrows)
            y_sdp = np.zeros(Nrows)
            z_sdp = np.zeros(Nrows)
            nn = 0
            for row in range(2, Nrows+1):
                t = Time(sheet[col_time+str(row)].value.encode('ascii')) - UTC_OFFSET
                x = sheet[col_SDP_X+str(row)].value
                y = sheet[col_SDP_Y+str(row)].value
                z = sheet[col_SDP_Z+str(row)].value
                # exclude NaNs and 0 for x, y, z.
                if np.isnan(x) or np.isnan(y) or np.isnan(z) or x==0 or y==0 or z==0: continue
                mjd_sdp[nn] = t.mjd  
                x_sdp[nn] = x
                y_sdp[nn] = y
                z_sdp[nn] = z
                timeList.append(t)
                nn += 1

            self.t = timeList
            self.mjd = mjd_sdp[:nn]
            self.x = x_sdp[:nn]
            self.y = y_sdp[:nn]
            self.z = z_sdp[:nn]
            self.pressure = 0.
            self.temperature = 0.
        else:
            [self.mjd, self.ra, self.dec, self.az, self.alt] = readListFile(posFile, self.fields)
             

    def xyz2azalt(self):
        """
        convert x, y, z to az, alt
        Return: az and alt in degrees
        """
        R_ca = np.sqrt(self.x*self.x+self.y*self.y+self.z*self.z)
        az0 = np.degrees(np.arctan2(self.x, self.y)) % 360.
        alt = 90. - np.degrees(np.arccos(-self.z/R_ca))
        az = np.copy(az0)

        # +180 if in the range [0,180.]
        sel1 = az0 >= 0.
        sel2 = az0 <= 180.
        sel = sel1 * sel2
        az[sel] += 180.

        # -180 if in the range (180, 360]
        sel1 = az0 > 180.
        sel2 = az0 <= 360.
        sel = sel1 * sel2
        az[sel] -= 180.

        self.az = az
        self.alt = alt


    def azalt2eq(self):
        """
        Convert az, alt to ra, dec
        Return: ra and dec in degrees
        """
        frame = AltAz(obstime=self.t, location=FAST_loc, pressure=self.pressure, temperature=self.temperature)
        c = SkyCoord(alt = self.alt*u.deg, az = self.az*u.deg, frame = frame)
        c_new = c.transform_to('fk5')
        self.ra = c_new.ra.deg
        self.dec = c_new.dec.deg


    def plotRADec(self,output=None):
        fig = plt.figure()
        ax = fig.add_subplot(1,1,1)
        ax.plot(self.ra, self.dec) 
        ax.set_xlabel('RA')
        ax.set_ylabel('Dec')
        if output==None: plt.show()
        else: plt.savefig(output)


    def runsave(self):
        """
        output as csv file
        """
        self.xyz2azalt()
        self.azalt2eq()
        allList = [self.mjd, self.ra, self.dec, self.az, self.alt]
        saveListFile(allList, self.fields, self.output)


class Scan:
    def __init__(self, ObsFile, fileNr, scanDir, t_track, ra_track, dec_track, RA_c, Dec_c, BeamNr, RotAngle, numSkip=5, extent=1.,outbase=None, Chan=42000):
        # read in all the data
        print("Reading all the data ....")
        for i in range(len(fileNr)):
            data = pf.getdata(ObsFile.replace('XXX',fileNr[i]))
            if i==0: 
                # observation time in mjd
                self.ObsTime = data['UTOBS'] 
                # ObsData is a 3D matrix, and its shape is (477, 65536, 4), 
                # corresponding to (row/or time sequence, frequency channels, Stokes).
                self.ObsData = data['DATA'][:,Chan,:]
                self.freq0 = data['FREQ'][0]
                self.bw = data['CHAN_BW'][0]
            else:
                self.ObsTime = np.concatenate((self.ObsTime, data['UTOBS']),axis=0)
                self.ObsData = np.concatenate((self.ObsData, data['DATA'][:,Chan,:]),axis=0)

        self.RA_s = RA_c
        self.Dec_s = Dec_c
        self.fghz = (self.freq0 + Chan * self.bw)/1.e3
        if outbase == None:
            self.outbase = RegEx.findall(r"M\d\d", ObsFile)[0]
        else:
            self.outbase = outbase

        ind1 = np.where(self.ObsTime >= t_track[0])[0][0]
        ind2 = np.where(self.ObsTime <= t_track[-1])[0][-1]

        self.ObsTime = self.ObsTime[ind1:ind2+1]
        self.ObsData = self.ObsData[ind1:ind2+1]

        print("Interpolation ...")
        f1 = interpolate.interp1d(t_track, ra_track, kind='linear')
        f2 = interpolate.interp1d(t_track, dec_track, kind='linear')

        self.ra = f1(self.ObsTime)
        self.dec = f2(self.ObsTime)

        """
        fig = plt.figure()
        ax = fig.add_subplot(1,1,1)
        #ax.scatter(self.ra, self.dec, c=self.ObsTime)
        ax.plot(self.ra[1:] - self.ra[:-1])
        plt.show()
        """

        coords = SkyCoord(ra = self.ra * u.deg, dec = self.dec * u.deg, frame='fk5')
        self.parangle = parallactic_angle(self.ObsTime, coords, FAST_loc)
        ## Check parallactic angle calculations
        #parangle_tt = parallactic_angle_sla(self.ObsTime, coords, FAST_loc)
        #fig = plt.figure()
        #ax = fig.add_subplot(1,1,1)
        #ax.plot(np.degrees(parangle_tt), np.degrees(abs(self.parangle-parangle_tt))*3600., 'o')
        #ax.set_xlabel('Parallactic angle (sla)')
        #ax.set_ylabel('Parallactic angle difference (arcsec)')
        #plt.show()

        shape = self.ObsData.shape
        self.NumStokes = shape[1]
        self.scanDir = scanDir

        if self.scanDir == 'RA':
            ttt = self.ra[1:] - self.ra[:-1]
        else:
            ttt = self.dec[1:] - self.dec[:-1]

        self.raList = []
        self.decList = []
        self.dataList = []
        self.parangleList = []

        for i in range(1,ttt.size):
            if i == 1: N1 = 0
            if ttt[i]*ttt[i-1] <= 0 and (i-N1)>numSkip:
                N2 = i
                self.raList.append(self.ra[N1:N2])
                self.decList.append(self.dec[N1:N2])
                self.dataList.append(self.ObsData[N1:N2])
                self.parangleList.append(self.parangle[N1:N2])
                N1 = N2
            if i == ttt.size-1 and (i-N1)>numSkip: N2 = ttt.size

        self.nscan = len(self.decList)
        print("Number of scans: ", self.nscan)

        if int(BeamNr[1:]) > 1:
            ss = BeamNr
            for i in range(self.nscan):
                dec_t = np.radians(self.decList[i])
                ra_t = np.radians(self.raList[i])
                dec_t = dec_t + BeamOffsetDict[ss][1] * BeamOffsetDict[ss][0].rad * np.sin((RotAngle + BeamOffsetDict[ss][3]).rad)
                ra_t = ra_t + BeamOffsetDict[ss][2] * BeamOffsetDict[ss][0].rad * np.cos((RotAngle + BeamOffsetDict[ss][3]).rad)/np.cos(dec_t)
                self.decList[i] = np.degrees(dec_t)
                self.raList[i] = np.degrees(ra_t)

        self.ext = extent
        if scanDir == 'RA':
            self.ra_c = RA_c
            self.dec_all = np.zeros(self.nscan)
            for i in range(self.nscan): self.dec_all[i] = np.mean(self.decList[i])
            if self.nscan == 1: self.cdelt = 1./60.
            else: self.cdelt = abs(np.median(self.dec_all[1:]-self.dec_all[:-1]))
            self.dec_c = self.dec_all[self.nscan//2] 
            self.ra_all = np.arange(self.ra_c+self.ext/2., self.ra_c-self.ext/2.-self.cdelt, -self.cdelt)
        else:
            self.dec_c = Dec_c
            self.ra_all = np.zeros(self.nscan)
            for i in range(self.nscan): self.ra_all[i] = np.mean(self.raList[i])
            self.cdelt = abs(np.median(self.ra_all[1:]-self.ra_all[:-1]))
            self.ra_c = self.ra_all[self.nscan//2] 
            self.dec_all = np.arange(self.dec_c-self.ext/2., self.dec_c+self.ext/2.+self.cdelt, self.cdelt)

        self.hdr = pf.Header({ 'SIMPLE':   True,
                               'BITPIX':    -32, 
                               'NAXIS':       2,
                               'NAXIS1':  self.ra_all.size,
                               'NAXIS2':  self.dec_all.size,
                               'CDELT1':  -self.cdelt,
                               'CRVAL1':  self.ra_c,
                               'CRPIX1':  (self.ra_c - self.ra_all[0])/(-self.cdelt) + 1.,
                               'CTYPE1': 'RA---CAR' ,
                               'CDELT2':   self.cdelt,
                               'CRPIX2':  (0.0 - self.dec_all[0]) / self.cdelt + 1. ,
                               'CRVAL2':   0.0,
                               'CTYPE2': 'DEC--CAR' ,
                               'EXTEND':True
                             })
        self.hdr['HISTORY'] = 'Test ...'
        self.hdr['COMMENT'] = 'Test ...'

        
    def WaterfallPlot(self, signal='ON'):
        # Waterfall plot
        fig = plt.figure()
        for i in range(self.NumStokes): 
            ax = fig.add_subplot(2, 2, i+1)
            if signal == 'ON':
                # ON  (discarded data of the 1st second)
                tmpdata = self.ObsData[2::2,:,i]
            elif signal == 'OFF':
                # OFF
                tmpdata = self.ObsData[1::2,:,i]
            elif signal == 'CAL':
                # Cal or Ref signal
                tmpdata = self.ObsData[2::2,:,i] - self.ObsData[1::2,:,i]
            else:
                # All
                tmpdata = self.ObsData[:,:,i]
            vmax = np.percentile(tmpdata,80)
            vmin = np.percentile(tmpdata,5)
            ax.imshow(tmpdata,origin='lower',aspect='auto',vmin=vmin, vmax=vmax)
        plt.show()


    def DisplayChannel(self):
        # for individual channel
        fig = plt.figure()
        for i in range(self.NumStokes): 
            Ton = self.ObsData[2::2,i]
            Toff = self.ObsData[1::2,i]
            T = Ton - Toff[:-1]
            ax = fig.add_subplot(2,2,i+1)
            #ax.plot(Ton,label='ON')
            #ax.plot(Toff,label='OFF')
            ax.plot(T,label='OFF')
            if i==0: ax.legend()
        plt.show()


    def CalWithInjectedNoise(self, Tcal=[10., 10.], check=False, checkMap=True, subBaseline=True, interp='expsinc', signQ=1.0, signU=1.0, para_corr=False, RotAngle=Angle(0.*u.deg), baseFrac='all', baselineOrder=1):
        # separate Sig and Cal
        N_r_p = 11 
        N_c_p = 2

        self.I_all = np.zeros((self.dec_all.size, self.ra_all.size))
        self.Q_all = np.zeros((self.dec_all.size, self.ra_all.size))
        self.U_all = np.zeros((self.dec_all.size, self.ra_all.size))
        self.V_all = np.zeros((self.dec_all.size, self.ra_all.size))

        for jj in range(self.nscan):  

            if check: fig = plt.figure()

            sig_all = []
            ref_all = []
            sign = 1.
            for i in range(self.NumStokes): 
                if check:
                    ax1 = fig.add_subplot(N_r_p,N_c_p,i+1)
                    ax2 = fig.add_subplot(N_r_p,N_c_p,i+5)
                Toff = self.dataList[jj][::2,i]
                Ton  = self.dataList[jj][1::2,i]

                if i==0: nsize = min(Toff.size, Ton.size)
                sig = Ton[:nsize] + Toff[:nsize]
                ref = Ton[:nsize] - Toff[:nsize]
                if i==0 and ref[ref>0].size - ref[ref<0].size < 0: sign = -1.

                ref *= sign

                # linear fit of ref signal
                p1 = models.Polynomial1D(1)
                pfit = fitting.LinearLSQFitter()
                sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                x = np.arange(ref.size)
                new_model, mask = sigma_clip_fit(p1, x, ref)
                ref_fit = new_model(x)
                if check:
                    ax1.plot(x, ref)
                    ax1.plot(x, ref_fit)
                    ax2.plot(x, sig)

                sig_all.append((sig - ref_fit) / 2.)
                ref_all.append(ref_fit * 1.)

            parangle = (self.parangleList[jj][::2][:nsize]+self.parangleList[jj][1::2][:nsize])/2.

            #   I' = | 1        scale_IQ |  I
            #   Q' = | scale_IQ    1     |  Q
            I_ref = (ref_all[0] + ref_all[1])/2.
            Q_ref = (ref_all[0] - ref_all[1])/2.
            U_ref = ref_all[2]*1.
            V_ref = ref_all[3]*1.
            scale_IQ = Q_ref / I_ref 
            ang_inst = np.arctan2(V_ref, U_ref) 
        
            I_ref_c = (I_ref - scale_IQ * Q_ref) / (1 - scale_IQ*scale_IQ)
            Q_ref_c = (Q_ref - scale_IQ * I_ref) / (1 - scale_IQ*scale_IQ)
            U_ref_c = U_ref * np.cos(ang_inst) + V_ref * np.sin(ang_inst)
            V_ref_c = V_ref * np.cos(ang_inst) - U_ref * np.sin(ang_inst)

            f_U = I_ref_c / U_ref_c
  
            scale_T = np.sqrt(Tcal[0]*Tcal[1]) / np.sqrt(ref_all[0]*ref_all[1]) 

            I_ref_c *= scale_T
            Q_ref_c *= scale_T
            U_ref_c *= scale_T
            V_ref_c *= scale_T

            I_sig = (sig_all[0] + sig_all[1])/2.
            Q_sig = (sig_all[0] - sig_all[1])/2.
            U_sig = sig_all[2]*1.
            V_sig = sig_all[3]*1.

            I_sig_c = (I_sig - scale_IQ * Q_sig) / (1 - scale_IQ*scale_IQ)
            Q_sig_c = (Q_sig - scale_IQ * I_sig) / (1 - scale_IQ*scale_IQ)
            U_sig_c = U_sig * np.cos(ang_inst) + V_sig * np.sin(ang_inst)
            V_sig_c = V_sig * np.cos(ang_inst) - U_sig * np.sin(ang_inst)

            U_sig_c *= f_U

            I_sig_c *= scale_T
            Q_sig_c = Q_sig_c * scale_T * signQ
            U_sig_c = U_sig_c * scale_T * signU
            V_sig_c *= scale_T

            # correction for feed rotation
            Q_sig_c1 = Q_sig_c * np.cos(2.*RotAngle).value + U_sig_c * np.sin(2.*RotAngle).value
            U_sig_c1 = -Q_sig_c * np.sin(2.*RotAngle).value + U_sig_c * np.cos(2.*RotAngle).value
            Q_sig_c = Q_sig_c1 * 1.
            U_sig_c = U_sig_c1 * 1.

            # correct for parangle
            if para_corr:
                Q_sig_c1 = Q_sig_c * np.cos(2.*parangle) - U_sig_c * np.sin(2.*parangle) 
                U_sig_c1 = Q_sig_c * np.sin(2.*parangle) + U_sig_c * np.cos(2.*parangle)
                Q_sig_c = Q_sig_c1 * 1.
                U_sig_c = U_sig_c1 * 1.

            if self.scanDir == "RA":
                xpos = (self.raList[jj][::2][:nsize]+self.raList[jj][1::2][:nsize])/2.
                rPos = self.ra_all
            else:
                xpos = (self.decList[jj][::2][:nsize]+self.decList[jj][1::2][:nsize])/2.
                rPos = self.dec_all

            if interp == "linear":
                fI = interpolate.interp1d(xpos, I_sig_c, fill_value=0., bounds_error=False)
                fQ = interpolate.interp1d(xpos, Q_sig_c, fill_value=0., bounds_error=False)
                fU = interpolate.interp1d(xpos, U_sig_c, fill_value=0., bounds_error=False)
                fV = interpolate.interp1d(xpos, V_sig_c, fill_value=0., bounds_error=False)
                y_fitI_all = fI(rPos)
                y_fitQ_all = fQ(rPos)
                y_fitU_all = fU(rPos)
                y_fitV_all = fV(rPos)
            elif interp == "expsinc":
                y_fitI_all = expsinc(xpos, I_sig_c, rPos, DIAM, self.fghz)
                y_fitQ_all = expsinc(xpos, Q_sig_c, rPos, DIAM, self.fghz)
                y_fitU_all = expsinc(xpos, U_sig_c, rPos, DIAM, self.fghz)
                y_fitV_all = expsinc(xpos, V_sig_c, rPos, DIAM, self.fghz)

            if subBaseline:
                print(f'base frac: {baseFrac}')
                if baseFrac == 'edge':
                    ind_ttt = np.arange(rPos.size)[y_fitI_all>0.]
                    ind1 = ind_ttt[0]
                    ind2 = ind1 + int(ind_ttt.size*0.15)
                    ind4 = ind_ttt[-1]+1
                    ind3 = ind4 - int(ind_ttt.size*0.15)
                    x_fit1 = np.concatenate((rPos[ind1:ind2], rPos[ind3:ind4]))
                    y_fitI = np.concatenate((y_fitI_all[ind1:ind2], y_fitI_all[ind3:ind4]))
                    y_fitQ = np.concatenate((y_fitQ_all[ind1:ind2], y_fitQ_all[ind3:ind4]))
                    y_fitU = np.concatenate((y_fitU_all[ind1:ind2], y_fitU_all[ind3:ind4]))
                    y_fitV = np.concatenate((y_fitV_all[ind1:ind2], y_fitV_all[ind3:ind4]))

                    p1 = models.Polynomial1D(baselineOrder)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pI, mask_I = sigma_clip_fit(p1, x_fit1, y_fitI)
                    #model_pI = pfit(p1, x_fit1, y_fitI)

                    p1 = models.Polynomial1D(baselineOrder)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pQ, mask_Q = sigma_clip_fit(p1, x_fit1, y_fitQ)
                    #model_pQ = pfit(p1, x_fit1, y_fitQ)

                    p1 = models.Polynomial1D(baselineOrder)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pU, mask_U = sigma_clip_fit(p1, x_fit1, y_fitU)
                    #model_pU = pfit(p1, x_fit1, y_fitU)

                    p1 = models.Polynomial1D(baselineOrder)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pV, mask_V = sigma_clip_fit(p1, x_fit1, y_fitV)
                    #model_pV = pfit(p1, x_fit1, y_fitV)

                    self.I_all[jj,ind1:ind4] = y_fitI_all[ind1:ind4] - model_pI(rPos[ind1:ind4])
                    self.Q_all[jj,ind1:ind4] = y_fitQ_all[ind1:ind4] - model_pQ(rPos[ind1:ind4])
                    self.U_all[jj,ind1:ind4] = y_fitU_all[ind1:ind4] - model_pU(rPos[ind1:ind4])
                    self.V_all[jj,ind1:ind4] = y_fitV_all[ind1:ind4] - model_pV(rPos[ind1:ind4])

                elif baseFrac == 'double-edge':
                    
                    ind_ttt = np.arange(rPos.size)[y_fitI_all>0.]
                    ind1 = int(ind_ttt.size*0.1)
                    ind2 = ind1 + int(ind_ttt.size*0.15)
                    ind4 = int(ind_ttt.size*0.9)
                    ind3 = ind4 - int(ind_ttt.size*0.15)
                    print(ind1,ind2,ind3,ind4)
                    x_fit1 = np.concatenate((rPos[ind1:ind2], rPos[ind3:ind4]))
                    y_fitI = np.concatenate((y_fitI_all[ind1:ind2], y_fitI_all[ind3:ind4]))
                    y_fitQ = np.concatenate((y_fitQ_all[ind1:ind2], y_fitQ_all[ind3:ind4]))
                    y_fitU = np.concatenate((y_fitU_all[ind1:ind2], y_fitU_all[ind3:ind4]))
                    y_fitV = np.concatenate((y_fitV_all[ind1:ind2], y_fitV_all[ind3:ind4]))

                    p1 = models.Polynomial1D(baselineOrder)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pI, mask_I = sigma_clip_fit(p1, x_fit1, y_fitI)
                    #model_pI = pfit(p1, x_fit1, y_fitI)

                    p1 = models.Polynomial1D(baselineOrder)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pQ, mask_Q = sigma_clip_fit(p1, x_fit1, y_fitQ)
                    #model_pQ = pfit(p1, x_fit1, y_fitQ)

                    p1 = models.Polynomial1D(baselineOrder)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pU, mask_U = sigma_clip_fit(p1, x_fit1, y_fitU)
                    #model_pU = pfit(p1, x_fit1, y_fitU)

                    p1 = models.Polynomial1D(baselineOrder)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pV, mask_V = sigma_clip_fit(p1, x_fit1, y_fitV)
                    #model_pV = pfit(p1, x_fit1, y_fitV)

                    self.I_all[jj,ind1:ind4] = y_fitI_all[ind1:ind4] - model_pI(rPos[ind1:ind4])
                    self.Q_all[jj,ind1:ind4] = y_fitQ_all[ind1:ind4] - model_pQ(rPos[ind1:ind4])
                    self.U_all[jj,ind1:ind4] = y_fitU_all[ind1:ind4] - model_pU(rPos[ind1:ind4])
                    self.V_all[jj,ind1:ind4] = y_fitV_all[ind1:ind4] - model_pV(rPos[ind1:ind4])


                elif baseFrac == 'all':
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pI, mask_I = sigma_clip_fit(p1, rPos, y_fitI_all)
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pQ, mask_Q = sigma_clip_fit(p1, rPos, y_fitQ_all)
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pU, mask_U = sigma_clip_fit(p1, rPos, y_fitU_all)
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pV, mask_V = sigma_clip_fit(p1, rPos, y_fitV_all)
                    self.I_all[jj] = y_fitI_all - model_pI(rPos)
                    self.Q_all[jj] = y_fitQ_all - model_pQ(rPos)
                    self.U_all[jj] = y_fitU_all - model_pU(rPos)
                    self.V_all[jj] = y_fitV_all - model_pV(rPos)

                elif baseFrac == 'arpls':
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pI, mask_I = sigma_clip_fit(p1, rPos, y_fitI_all)
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pQ, mask_Q = sigma_clip_fit(p1, rPos, y_fitQ_all)
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pU, mask_U = sigma_clip_fit(p1, rPos, y_fitU_all)
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pV, mask_V = sigma_clip_fit(p1, rPos, y_fitV_all)
                    self.I_all[jj] = y_fitI_all - model_pI(rPos)
                    self.Q_all[jj] = y_fitQ_all - model_pQ(rPos)
                    self.U_all[jj] = y_fitU_all - model_pU(rPos)
                    self.V_all[jj] = y_fitV_all - model_pV(rPos)

            else:
                self.I_all[jj,:] = y_fitI_all
                self.Q_all[jj,:] = y_fitQ_all
                self.U_all[jj,:] = y_fitU_all
                self.V_all[jj,:] = y_fitV_all

            if check:
                ax9 = fig.add_subplot(N_r_p, N_c_p, 9)
                ax9.plot(x, I_ref_c)
                ax10 = fig.add_subplot(N_r_p, N_c_p, 10)
                ax10.plot(x, Q_ref_c)
                ax11 = fig.add_subplot(N_r_p, N_c_p, 11)
                #ax11.plot(x, scale_IQ)
                ax11.plot(x, U_ref_c)
                ax12 = fig.add_subplot(N_r_p, N_c_p, 12)
                #ax12.plot(x, ang_inst)
                ax12.plot(x, V_ref_c)
                ax13 = fig.add_subplot(N_r_p, N_c_p, 13)
                ax13.plot(x, I_sig_c)
                ax14 = fig.add_subplot(N_r_p, N_c_p, 14)
                ax14.plot(x, Q_sig_c)
                ax15 = fig.add_subplot(N_r_p, N_c_p, 15)
                ax15.plot(x, U_sig_c)
                ax16 = fig.add_subplot(N_r_p, N_c_p, 16)
                ax16.plot(x, V_sig_c)
                ax17 = fig.add_subplot(N_r_p, N_c_p, 17)
                ax17.plot(x, parangle)
                if self.scanDir == "RA":
                    ax19 = fig.add_subplot(N_r_p, N_c_p, 19)
                    ax19.plot(self.ra_all, self.I_all[jj])
                    ax20 = fig.add_subplot(N_r_p, N_c_p, 20)
                    ax20.plot(self.ra_all, self.Q_all[jj])
                    ax21 = fig.add_subplot(N_r_p, N_c_p, 21)
                    ax21.plot(self.ra_all, self.U_all[jj])
                    ax22 = fig.add_subplot(N_r_p, N_c_p, 22)
                    ax22.plot(self.ra_all, self.V_all[jj])
                else:
                    ax19 = fig.add_subplot(N_r_p, N_c_p, 19)
                    ax19.plot(self.dec_all, self.I_all[:,jj])
                    ax20 = fig.add_subplot(N_r_p, N_c_p, 20)
                    ax20.plot(self.dec_all, self.Q_all[:,jj])
                    ax21 = fig.add_subplot(N_r_p, N_c_p, 21)
                    ax21.plot(self.dec_all, self.U_all[:,jj])
                    ax22 = fig.add_subplot(N_r_p, N_c_p, 22)
                    ax22.plot(self.dec_all, self.V_all[:,jj])
                plt.show()

        self.hdr['OBJECT'] = 'I'
        pf.writeto(self.outbase+"_I.fits", self.I_all, self.hdr, overwrite=True)
        self.hdr['OBJECT'] = 'Q'
        pf.writeto(self.outbase+"_Q.fits", self.Q_all, self.hdr, overwrite=True)
        self.hdr['OBJECT'] = 'U'
        pf.writeto(self.outbase+"_U.fits", self.U_all, self.hdr, overwrite=True)
        self.hdr['OBJECT'] = 'V'
        pf.writeto(self.outbase+"_V.fits", self.V_all, self.hdr, overwrite=True)

        if checkMap:
            fig = plt.figure()
            ax1 = fig.add_subplot(2,2,1)
            fI_data = maputils.FITSimage(externalheader=self.hdr, externaldata=self.I_all)
            annim1 = fI_data.Annotatedimage(ax1)
            annim1.Image(interpolation='spline36')
            grat1 = annim1.Graticule()
            grat1.setp_axislabel("bottom", visible=False)
            grat1.setp_ticklabel(wcsaxis=0, visible=False)
            annim1.plot()

            ax2 = fig.add_subplot(2,2,2)
            fQ_data = maputils.FITSimage(externalheader=self.hdr, externaldata=self.Q_all)
            annim2 = fQ_data.Annotatedimage(ax2)
            annim2.Image(interpolation='spline36')
            grat2 = annim2.Graticule()
            grat2.setp_axislabel("bottom", visible=False)
            grat2.setp_axislabel("left", visible=False)
            grat2.setp_ticklabel(wcsaxis=0, visible=False)
            grat2.setp_ticklabel(wcsaxis=1, visible=False)
            annim2.plot()

            ax3 = fig.add_subplot(2,2,3)
            fU_data = maputils.FITSimage(externalheader=self.hdr, externaldata=self.U_all)
            annim3 = fU_data.Annotatedimage(ax3)
            annim3.Image(interpolation='spline36')
            grat3 = annim3.Graticule()
            annim3.plot()

            ax4 = fig.add_subplot(2,2,4)
            fV_data = maputils.FITSimage(externalheader=self.hdr, externaldata=self.V_all)
            annim4 = fV_data.Annotatedimage(ax4)
            annim4.Image(interpolation='spline36')
            grat4 = annim4.Graticule()
            grat4.setp_axislabel("left", visible=False)
            grat4.setp_ticklabel(wcsaxis=1, visible=False)
            annim4.plot()
            fig.savefig(self.outbase+".png")
            plt.close()


    def GetPeakValues(self, x=[np.zeros(0), np.zeros(0)], radec = None, mode='peak_p', ave_num=1, guess_range=5., signals=None, check=True, figname=None):
        # get the gain factor by ffitting calibrator
     
        if signals == None:
            signals = [self.I_all, self.Q_all, self.U_all, self.V_all]
        if x[0].size == 0:
            XI, YI = np.meshgrid(self.ra_all, self.dec_all)
        else:
            XI, YI = np.meshgrid(x[0], x[1])

        if radec == None:
            ra_c = self.RA_s
            dec_c = self.Dec_s 

        self.sig_obs = []

        for i in range(len(signals)):
            sig = signals[i]
            M, N = sig.shape

            if i==0:
                sign_f = 1.

                tt = abs(XI[0] - ra_c)
                ix = np.argmin(tt)
                tt = abs(YI[:,0] - dec_c)
                iy = np.argmin(tt)

                if sig[iy, ix] < 0.: sign_f = -1.
                y_f = sig * sign_f

                p_init = models.Polynomial2D(1)
                p_init.c0_0 = np.median(y_f)

                ytt = y_f - np.median(y_f)
                amp = ytt[iy, ix]

                sel_t1 = (abs(XI - ra_c) < guess_range/60.)
                sel_t2 = (abs(YI - dec_c) < guess_range/60.)
                sel_t = sel_t1 & sel_t2

                x_mean = np.mean(XI[sel_t]*ytt[sel_t])/np.mean(ytt[sel_t])
                y_mean = np.mean(YI[sel_t]*ytt[sel_t])/np.mean(ytt[sel_t])
                x_std = np.sqrt(abs(np.sum((XI[sel_t]-x_mean)*(XI[sel_t]-x_mean)*(ytt[sel_t])))/np.sum(ytt[sel_t]))
                y_std = np.sqrt(abs(np.sum((YI[sel_t]-y_mean)*(YI[sel_t]-y_mean)*(ytt[sel_t])))/np.sum(ytt[sel_t]))
                g_init = models.Gaussian2D(amplitude=amp, x_mean=x_mean, y_mean=y_mean, x_stddev=x_std, y_stddev=y_std)
                
                pg = p_init + g_init

                fit = fitting.LevMarLSQFitter()
                #fit = fitting.SLSQPLSQFitter()

                fitted_model = fit(pg, XI, YI, y_f)
                ierr = fit.fit_info['ierr']
                Intensity = fitted_model[1].amplitude.value/sign_f
                x_beam = fitted_model[1].x_stddev.value*2.*np.sqrt(2.*np.log(2.)) 
                y_beam = fitted_model[1].y_stddev.value*2.*np.sqrt(2.*np.log(2.)) 
                theta = fitted_model[1].theta.value
                x_mean = fitted_model[1].x_mean.value
                y_mean = fitted_model[1].y_mean.value
                self.sig_obs.append([Intensity, x_beam, y_beam, theta, x_mean, y_mean])
                if check == True and ierr != 0:
                    if i==0: fig = plt.figure(figsize=(10,8))
                    ax1 = fig.add_subplot(3,2,i+1)
                    ax1.imshow(sig, vmin = np.percentile(sig,2.), vmax=np.percentile(sig, 98.))
                    ax2 = fig.add_subplot(3,2,i+5)
                    tttt = sig - fitted_model(XI, YI)/sign_f 
                    ax2.imshow(tttt, vmin = np.percentile(tttt, 2.), vmax=np.percentile(tttt,98.))
                xc = x_mean
                yc = y_mean
                if mode == 'peak_p':
                    d = abs(XI[0][0]-XI[0][1]) 
                    x_ind = (abs(XI[0] - xc) < (2.*d * ave_num))
                    y_ind = (abs(YI[:,0] - yc) < (2.*d * ave_num))
                    IX, IY = np.meshgrid(x_ind, y_ind)
                    ii = IX & IY
                    INDx, INDy = np.meshgrid(np.arange(N), np.arange(M))
                    xp_1 = INDx[ii]
                    yp_1 = INDy[ii]
                    zp = sig[ii]
                    i_p = zp.argmax()
                    xp = xp_1[i_p]
                    yp = yp_1[i_p]
            else:
                if mode == 'ave':
                    d = abs(XI[0][0]-XI[0][1]) 
                    x_ind = (abs(XI[0] - xc) < (d * ave_num))
                    y_ind = (abs(YI[:,0] - yc) < (d * ave_num))
                    IX, IY = np.meshgrid(x_ind, y_ind)
                    ii = IX & IY
                    INDx, INDy = np.meshgrid(np.arange(N), np.arange(M))
                    self.sig_obs.append([np.mean(sig[ii]) - np.median(sig)])
                    xp = INDx[ii]
                    yp = INDy[ii]
                elif mode == 'peak_f':
                    tt = abs(XI[0] - xc)
                    ix = np.argmin(tt)
                    tt = abs(YI[:,0] - yc)
                    iy = np.argmin(tt)
                    self.sig_obs.append([np.mean(sig[iy, ix]) - np.median(sig)])
                    xp = np.array([ix])
                    yp = np.array([iy])
                elif mode == 'peak_p':
                    self.sig_obs.append([sig[yp, xp] - np.median(sig)])

                if check == True and ierr != 0:
                    ax1 = fig.add_subplot(3,2,i+1)
                    ax1.imshow(sig, vmin = np.percentile(sig,2.), vmax=np.percentile(sig, 98.))
                    ax1.plot(xp, yp, 'ro', alpha=0.6)

        ti = self.sig_obs[0][0]
        q = self.sig_obs[1][0]
        u = self.sig_obs[2][0]
        v = self.sig_obs[3][0]
        pi = np.sqrt(q*q + u*u)
        pc = pi*100./ti
        pa = np.degrees(0.5*np.arctan2(u, q))

        if check == True and ierr !=0:
            if figname == None: plt.show()
            else: 
                plt.savefig(figname)
                plt.close()

        return ti, q, u, v, pi, pa, pc, ierr



    def GetCalFactors_S(self, source='3C286', freq = None, signals = None):

        if signals == None:
            signals = []
            for i in range(len(self.sig_obs)):
                signals.append(self.sig_obs[i][0])

        if freq == None: freq = self.fghz

        I0 = flux_model(source, freq)
        pc, pa = pol_model(source, freq)
        Q0 = I0*pc*np.cos(2.*np.radians(pa))/100.
        U0 = I0*pc*np.sin(2.*np.radians(pa))/100.
        V0 = 0.

        f_I = I0 / signals[0]
        f_Q = Q0 / signals[1]
        f_U = U0 / signals[2]

        return f_I, f_Q, f_U


def makeMap(kyFile, posFile, infTemplate, fileNr, BeamList, RA_c, Dec_c, scanDir, good_chan_file = None, tcalFile=None, baseFrac='all', RotAngle=Angle(0.*u.deg), numSkip=20, extent=1., ChanList=[42000], interp = 'expsinc', signQ=-1.0, signU=1.0, para_corr=False, check=False, checkMap=True, subBaseline=True, doCombine=False, outFileBase='all', mode=['GetCal', '3c138', '3c138cal.dat']):
    if not os.path.exists(posFile):
        # Run for first time, calculate RA/DEC
        track = Track(posFile, kyFile=kyFile)
        track.runsave()
    else:
        # Second time, read positions from the file
        track = Track(posFile)
        #track.plotRADec(output='ttt1.png')

    if mode[0] == 'GetCal':
        source = mode[1]
        fout = open(mode[2], 'w')

    if tcalFile != None:
        tcals0 = np.fromfile(tcalFile, sep="\n")
        tcals = tcals0.reshape((tcals0.size//39, 39))

    if good_chan_file != None:
        tmp = np.fromfile(good_chan_file, sep="\n")
        tmp = tmp.reshape((tmp.size//2, 2))
        good_flag = tmp[:,1]

    for Chan in ChanList: 
        print("Channel: ", Chan)
        if good_chan_file != None:
            if good_flag[Chan] < 0.5:
                print("bad channel!")
                continue
        for i in BeamList:
            print("Beam: ",i)
            BeamNr = "M%02d" % i
            inf = infTemplate.replace('YYY', BeamNr[1:])
            outbase = outFileBase + '_chan_%s_b_%s' % (str(Chan), str(i))
            obs = Scan(inf, fileNr, scanDir, track.mjd, track.ra, track.dec, RA_c, Dec_c, BeamNr, RotAngle, numSkip=numSkip, extent=extent, outbase=outbase, Chan=Chan)

            try:
                if tcalFile != None:
                    freq_t = tcals[:,0]
                    tcal_xx = tcals[:,2*i-1]
                    tcal_yy = tcals[:,2*i]
                    freq0 = obs.fghz*1.e3
                    tcal_interp_xx = interpolate.interp1d(freq_t, tcal_xx, kind='linear')
                    tcal_interp_yy = interpolate.interp1d(freq_t, tcal_yy, kind='linear')
                    Tcal = [tcal_interp_xx(freq0), tcal_interp_yy(freq0)]
                else:
                    Tcal = [10., 10.]

                obs.CalWithInjectedNoise(Tcal = Tcal, check=check, checkMap=checkMap, subBaseline=subBaseline, interp=interp, signQ=signQ, signU=signU, para_corr=para_corr, RotAngle=RotAngle, baseFrac=baseFrac)
                if mode[0] == 'GetCal':
                    ti, q, u, v, pi, pa, pc, ierr = obs.GetPeakValues(check=True, figname=outbase+".peaks.png")
                    f_I, f_Q, f_U = obs.GetCalFactors_S(source=source)
                    ss = str(i)+" "+str(Chan)+" "+str(f_I)+" "+str(f_Q)+" "+str(f_U)+" "+str(ti)+" "+str(q)+" "+str(u)+" "+str(v)+" "+str(pi)+" "+str(pa)+" "+str(pc)+" "+str(obs.fghz)+" "+str(ierr)+"\n"
                    fout.write(ss)
            except:
                pass

        if doCombine:
            ra_c = RA_c
            dec_c = Dec_c
            ext = extent 
            cdelt = ext/60. # arcmin
            Num = int(ext/cdelt+1)
            if Num %2 == 0: Num += 1
            ra_0 = ra_c + cdelt*(Num//2)
            dec_0 = dec_c - cdelt*(Num//2)

            hdr = pf.Header({ 'SIMPLE':   True,
                              'BITPIX':    -32, 
                              'NAXIS':       2,
                              'NAXIS1':  Num,
                              'NAXIS2':  Num,
                              'CDELT1':  -cdelt,
                              'CRVAL1':  ra_c,
                              'CRPIX1':  (ra_c - ra_0)/(-cdelt) + 1.,
                              'CTYPE1': 'RA---CAR' ,
                              'CDELT2':   cdelt,
                              'CRPIX2':  (0.0 - dec_0) / cdelt + 1. ,
                              'CRVAL2':   0.0,
                              'CTYPE2': 'DEC--CAR' ,
                              'EXTEND':True
                            })
            hdr['HISTORY'] = 'Test ...'
            hdr['COMMENT'] = 'Test ...'

    
            output = outFileBase + '_%d_I.fits' % Chan
            combineBeams(hdr, output, stokes='I')
            output = outFileBase + '_%d_Q.fits' % Chan
            combineBeams(hdr, output, stokes='Q')
            output = outFileBase + '_%d_U.fits' % Chan
            combineBeams(hdr, output, stokes='U')
            output = outFileBase + '_%d_V.fits' % Chan
            combineBeams(hdr, output, stokes='V')
 
            udata = pf.getdata(outFileBase + '_%d_U.fits' % Chan)
            qdata = pf.getdata(outFileBase + '_%d_Q.fits' % Chan)
            pdata = np.sqrt(udata*udata+qdata*qdata)
            pf.writeto(outFileBase + '_P.fits', pdata, hdr, overwrite=True)

            fig = plt.figure()
            ax1 = fig.add_subplot(2,2,1)
            plotImage(ax1, outFileBase + '_%d_I.fits' % Chan, glvisible=False, gbvisible=True, gltickvisible=False, gbtickvisible=True)
            ax2 = fig.add_subplot(2,2,2)
            plotImage(ax2, outFileBase + '_%d_Q.fits' % Chan, glvisible=False, gbvisible=False, gltickvisible=False, gbtickvisible=False)
            ax3 = fig.add_subplot(2,2,3)
            plotImage(ax3, outFileBase + '_%d_U.fits' % Chan, glvisible=True, gbvisible=True, gltickvisible=True, gbtickvisible=True)
            ax4 = fig.add_subplot(2,2,4)
            plotImage(ax4, outFileBase + '_%d_P.fits' % Chan, glvisible=True, gbvisible=False, gltickvisible=True, gbtickvisible=False)
            plt.tight_layout()
            #plt.show()
            fig.savefig(outFileBase + '_%d_.png' % Chan)

    if mode[0] == 'GetCal': fout.close()


def make_tmp(file_region, data):
    print(f'the shape of orignal data is {data.shape}')
    reg = pyregion.open(file_region)
    data_mask = ~reg.get_mask(shape=data.shape)
    data_tmp = np.where(data_mask, np.nan, np.squeeze(data))
    print(f'the shape of template data is {data_tmp.shape}')
    # plt.imshow(data_tmp)
    # plt.show()
    # plt.close()
    return data_tmp

class ScanOne:
    def __init__(self, data, t_track, ra_track, dec_track, BeamNr, RotAngle, scanDir, numSkip=5, Chan=42000, ):
        # read in all the data

        # observation time in mjd
        self.ObsTime = data['UTOBS'] 
        # ObsData is a 3D matrix, and its shape is (477, 65536, 4), 
        # corresponding to (row/or time sequence, frequency channels, Stokes).
        self.ObsData = data['DATA'][:,Chan,:]
        self.freq0 = data['FREQ'][0]
        self.bw = data['CHAN_BW'][0]
        self.fghz = (self.freq0 + Chan * self.bw)/1.e3
        self.scanDir = scanDir

        ind1 = np.where(self.ObsTime >= t_track[0])[0][0]
        ind2 = np.where(self.ObsTime <= t_track[-1])[0][-1]
        ind1 += numSkip
        ind2 -= numSkip

        self.ObsTime = self.ObsTime[ind1:ind2+1]
        self.ObsData = self.ObsData[ind1:ind2+1]

        print("Interpolation ...")
        f1 = interpolate.interp1d(t_track, ra_track, kind='linear')
        f2 = interpolate.interp1d(t_track, dec_track, kind='linear')

        self.ra = f1(self.ObsTime)
        self.dec = f2(self.ObsTime)

        coords = SkyCoord(ra = self.ra * u.deg, dec = self.dec * u.deg, frame='fk5')
        self.parangle = parallactic_angle(self.ObsTime, coords, FAST_loc)

        shape = self.ObsData.shape
        self.NumStokes = shape[1]

        if int(BeamNr[1:]) > 1:
            ss = BeamNr
            dec_t = np.radians(self.dec)
            ra_t = np.radians(self.ra)
            dec_t = dec_t + BeamOffsetDict[ss][1] * BeamOffsetDict[ss][0].rad * np.sin((RotAngle + BeamOffsetDict[ss][3]).rad)
            ra_t = ra_t + BeamOffsetDict[ss][2] * BeamOffsetDict[ss][0].rad * np.cos((RotAngle + BeamOffsetDict[ss][3]).rad)/np.cos(dec_t)
            self.dec = np.degrees(dec_t)
            self.ra = np.degrees(ra_t)


        
    def WaterfallPlot(self, signal='ON'):
        # Waterfall plot
        fig = plt.figure()
        for i in range(self.NumStokes): 
            ax = fig.add_subplot(2, 2, i+1)
            if signal == 'ON':
                # ON  (discarded data of the 1st second)
                tmpdata = self.ObsData[2::2,:,i]
            elif signal == 'OFF':
                # OFF
                tmpdata = self.ObsData[1::2,:,i]
            elif signal == 'CAL':
                # Cal or Ref signal
                tmpdata = self.ObsData[2::2,:,i] - self.ObsData[1::2,:,i]
            else:
                # All
                tmpdata = self.ObsData[:,:,i]
            vmax = np.percentile(tmpdata,80)
            vmin = np.percentile(tmpdata,5)
            ax.imshow(tmpdata,origin='lower',aspect='auto',vmin=vmin, vmax=vmax)
        plt.show()


    def DisplayChannel(self):
        # for individual channel
        fig = plt.figure()
        for i in range(self.NumStokes): 
            Ton = self.ObsData[2::2,i]
            Toff = self.ObsData[1::2,i]
            T = Ton - Toff[:-1]
            ax = fig.add_subplot(2,2,i+1)
            ax.plot(Ton,label='ON')
            ax.plot(Toff,label='OFF')
            #ax.plot(T,label='OFF')
            if i==0: ax.legend()
        plt.show()


    def CalWithInjectedNoise(self, Tcal=[10., 10.], check=False, subBaseline=True, signQ=1.0, signU=1.0, para_corr=False, RotAngle=Angle(0.*u.deg), baseFrac='all',baselineOrder=1, edgeFrac=[0.15,0.15]):
        # separate Sig and Cal
        N_r_p = 9 
        N_c_p = 2

        if check: fig = plt.figure()


        sig_all = []
        ref_all = []
        sign = 1.
        for i in range(self.NumStokes): 
            if check:
                ax1 = fig.add_subplot(N_r_p,N_c_p,i+1)
                ax2 = fig.add_subplot(N_r_p,N_c_p,i+5)
            Toff = self.ObsData[::2,i]
            Ton  = self.ObsData[1::2,i]

            if i==0: nsize = min(Toff.size, Ton.size)
            sig = Ton[:nsize] + Toff[:nsize]
            ref = Ton[:nsize] - Toff[:nsize]
            if i==0 and ref[ref>0].size - ref[ref<0].size < 0: sign = -1.

            ref *= sign

            # linear fit of ref signal
            p1 = models.Polynomial1D(1)
            pfit = fitting.LinearLSQFitter()
            sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
            x = np.arange(ref.size)
            new_model, mask = sigma_clip_fit(p1, x, ref)
            ref_fit = new_model(x)
            if check:
                ax1.plot(x, ref)
                ax1.plot(x, ref_fit)
                ax2.plot(x, sig)

            sig_all.append((sig - ref_fit) / 2.)
            ref_all.append(ref_fit * 1.)
            if i == 0: sel_i1 = (ref_fit > 0.)
            if i == 1: sel_i2 = (ref_fit > 0.)

        parangle = (self.parangle[::2][:nsize]+self.parangle[1::2][:nsize])/2.
        ra0 = (self.ra[::2][:nsize]+self.ra[1::2][:nsize])/2.
        dec0 = (self.dec[::2][:nsize]+self.dec[1::2][:nsize])/2.
        self.ra = ra0
        self.dec = dec0

        sel_i1i2 = sel_i1 & sel_i2
        for i in range(self.NumStokes):
            ref_all[i] = ref_all[i][sel_i1i2]
            sig_all[i] = sig_all[i][sel_i1i2]
        self.ra = self.ra[sel_i1i2]
        self.dec = self.dec[sel_i1i2]
        parangle = parangle[sel_i1i2]


        #   I' = | 1        scale_IQ |  I
        #   Q' = | scale_IQ    1     |  Q
        I_ref = (ref_all[0] + ref_all[1])/2.
        Q_ref = (ref_all[0] - ref_all[1])/2.
        U_ref = ref_all[2]*1.
        V_ref = ref_all[3]*1.
        scale_IQ = Q_ref / I_ref 
        ang_inst = np.arctan2(V_ref, U_ref) 
        
        I_ref_c = (I_ref - scale_IQ * Q_ref) / (1 - scale_IQ*scale_IQ)
        Q_ref_c = (Q_ref - scale_IQ * I_ref) / (1 - scale_IQ*scale_IQ)
        U_ref_c = U_ref * np.cos(ang_inst) + V_ref * np.sin(ang_inst)
        V_ref_c = V_ref * np.cos(ang_inst) - U_ref * np.sin(ang_inst)

        f_U = I_ref_c / U_ref_c
  
        scale_T = np.sqrt(Tcal[0]*Tcal[1]) / np.sqrt(ref_all[0]*ref_all[1]) 

        I_ref_c *= scale_T
        Q_ref_c *= scale_T
        U_ref_c *= scale_T
        V_ref_c *= scale_T

        I_sig = (sig_all[0] + sig_all[1])/2.
        Q_sig = (sig_all[0] - sig_all[1])/2.
        U_sig = sig_all[2]*1.
        V_sig = sig_all[3]*1.

        I_sig_c = (I_sig - scale_IQ * Q_sig) / (1 - scale_IQ*scale_IQ)
        Q_sig_c = (Q_sig - scale_IQ * I_sig) / (1 - scale_IQ*scale_IQ)
        U_sig_c = U_sig * np.cos(ang_inst) + V_sig * np.sin(ang_inst)
        V_sig_c = V_sig * np.cos(ang_inst) - U_sig * np.sin(ang_inst)

        U_sig_c *= f_U

        I_sig_c *= scale_T
        Q_sig_c = Q_sig_c * scale_T * signQ
        U_sig_c = U_sig_c * scale_T * signU
        V_sig_c *= scale_T

        # correction for feed rotation
        Q_sig_c1 = Q_sig_c * np.cos(2.*RotAngle).value + U_sig_c * np.sin(2.*RotAngle).value
        U_sig_c1 = -Q_sig_c * np.sin(2.*RotAngle).value + U_sig_c * np.cos(2.*RotAngle).value
        Q_sig_c = Q_sig_c1 * 1.
        U_sig_c = U_sig_c1 * 1.

        # correct for parangle
        if para_corr:
            Q_sig_c1 = Q_sig_c * np.cos(2.*parangle) - U_sig_c * np.sin(2.*parangle) 
            U_sig_c1 = Q_sig_c * np.sin(2.*parangle) + U_sig_c * np.cos(2.*parangle)
            Q_sig_c = Q_sig_c1 * 1.
            U_sig_c = U_sig_c1 * 1.

        self.I_sig_c = I_sig_c
        self.Q_sig_c = Q_sig_c
        self.U_sig_c = U_sig_c
        self.V_sig_c = V_sig_c

        if subBaseline:
            if self.scanDir == 'RA':
                rPos = self.ra*1.
            else:
                rPos = self.dec*1.

            if baseFrac == 'edge':
                ind_ttt = np.arange(rPos.size)[I_sig_c>0.]
                if len(ind_ttt) == 0:
                    ind1 = 0
                    ind2 = rPos.size // 2 
                    ind3 = rPos.size // 2 
                    ind4 = rPos.size - 1 
                else:
                    ind1 = ind_ttt[0]
                    ind2 = ind1 + int(ind_ttt.size*edgeFrac[0])
                    ind4 = ind_ttt[-1]+1
                    ind3 = ind4 - int(ind_ttt.size*edgeFrac[1])


                x_fit1 = np.concatenate((rPos[ind1:ind2], rPos[ind3:ind4]))
                y_fitI = np.concatenate((I_sig_c[ind1:ind2], I_sig_c[ind3:ind4]))
                y_fitQ = np.concatenate((Q_sig_c[ind1:ind2], Q_sig_c[ind3:ind4]))
                y_fitU = np.concatenate((U_sig_c[ind1:ind2], U_sig_c[ind3:ind4]))
                y_fitV = np.concatenate((V_sig_c[ind1:ind2], V_sig_c[ind3:ind4]))

                p1 = models.Polynomial1D(baselineOrder)
                pfit = fitting.LinearLSQFitter()
                sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                model_pI, mask_I = sigma_clip_fit(p1, x_fit1, y_fitI)
                #model_pI = pfit(p1, x_fit1, y_fitI)

                p1 = models.Polynomial1D(baselineOrder)
                pfit = fitting.LinearLSQFitter()
                sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                model_pQ, mask_Q = sigma_clip_fit(p1, x_fit1, y_fitQ)
                #model_pQ = pfit(p1, x_fit1, y_fitQ)

                p1 = models.Polynomial1D(baselineOrder)
                pfit = fitting.LinearLSQFitter()
                sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                model_pU, mask_U = sigma_clip_fit(p1, x_fit1, y_fitU)
                #model_pU = pfit(p1, x_fit1, y_fitU)

                p1 = models.Polynomial1D(baselineOrder)
                pfit = fitting.LinearLSQFitter()
                sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                model_pV, mask_V = sigma_clip_fit(p1, x_fit1, y_fitV)
                #model_pV = pfit(p1, x_fit1, y_fitV)

                I_sig_c[ind1:ind4] = I_sig_c[ind1:ind4] - model_pI(rPos[ind1:ind4])
                Q_sig_c[ind1:ind4] = Q_sig_c[ind1:ind4] - model_pQ(rPos[ind1:ind4])
                U_sig_c[ind1:ind4] = U_sig_c[ind1:ind4] - model_pU(rPos[ind1:ind4])
                V_sig_c[ind1:ind4] = V_sig_c[ind1:ind4] - model_pV(rPos[ind1:ind4])
                self.ra = self.ra[ind1:ind4]
                self.dec = self.dec[ind1:ind4]
                self.I_sig_c = I_sig_c[ind1:ind4]
                self.Q_sig_c = Q_sig_c[ind1:ind4]
                self.U_sig_c = U_sig_c[ind1:ind4]
                self.V_sig_c = V_sig_c[ind1:ind4]

            elif baseFrac == 'double-edge':
                    
                ind_ttt = np.arange(rPos.size)[I_sig_c>0.]
                ind1 = int(ind_ttt.size*0.1)
                ind2 = ind1 + int(ind_ttt.size*0.15)
                ind4 = int(ind_ttt.size*0.9)
                ind3 = ind4 - int(ind_ttt.size*0.15)
                print(ind1,ind2,ind3,ind4)
                x_fit1 = np.concatenate((rPos[ind1:ind2], rPos[ind3:ind4]))
                y_fitI = np.concatenate((I_sig_c[ind1:ind2], I_sig_c[ind3:ind4]))
                y_fitQ = np.concatenate((Q_sig_c[ind1:ind2], Q_sig_c[ind3:ind4]))
                y_fitU = np.concatenate((U_sig_c[ind1:ind2], U_sig_c[ind3:ind4]))
                y_fitV = np.concatenate((V_sig_c[ind1:ind2], V_sig_c[ind3:ind4]))

                p1 = models.Polynomial1D(baselineOrder)
                pfit = fitting.LinearLSQFitter()
                sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                model_pI, mask_I = sigma_clip_fit(p1, x_fit1, y_fitI)
                #model_pI = pfit(p1, x_fit1, y_fitI)

                p1 = models.Polynomial1D(baselineOrder)
                pfit = fitting.LinearLSQFitter()
                sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                model_pQ, mask_Q = sigma_clip_fit(p1, x_fit1, y_fitQ)
                #model_pQ = pfit(p1, x_fit1, y_fitQ)

                p1 = models.Polynomial1D(baselineOrder)
                pfit = fitting.LinearLSQFitter()
                sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                model_pU, mask_U = sigma_clip_fit(p1, x_fit1, y_fitU)
                #model_pU = pfit(p1, x_fit1, y_fitU)

                p1 = models.Polynomial1D(baselineOrder)
                pfit = fitting.LinearLSQFitter()
                sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                model_pV, mask_V = sigma_clip_fit(p1, x_fit1, y_fitV)
                #model_pV = pfit(p1, x_fit1, y_fitV)

                I_sig_c[ind1:ind4] = I_sig_c[ind1:ind4] - model_pI(rPos[ind1:ind4])
                Q_sig_c[ind1:ind4] = Q_sig_c[ind1:ind4] - model_pQ(rPos[ind1:ind4])
                U_sig_c[ind1:ind4] = U_sig_c[ind1:ind4] - model_pU(rPos[ind1:ind4])
                V_sig_c[ind1:ind4] = V_sig_c[ind1:ind4] - model_pV(rPos[ind1:ind4])
                self.ra = self.ra[ind1:ind4]
                self.dec = self.dec[ind1:ind4]
                self.I_sig_c = I_sig_c[ind1:ind4]
                self.Q_sig_c = Q_sig_c[ind1:ind4]
                self.U_sig_c = U_sig_c[ind1:ind4]
                self.V_sig_c = V_sig_c[ind1:ind4]




            elif baseFrac == 'all':
                try:
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pI, mask_I = sigma_clip_fit(p1, rPos, I_sig_c)
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pQ, mask_Q = sigma_clip_fit(p1, rPos, Q_sig_c)
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pU, mask_U = sigma_clip_fit(p1, rPos, U_sig_c)
                    p1 = models.Polynomial1D(1)
                    pfit = fitting.LinearLSQFitter()
                    sigma_clip_fit = fitting.FittingWithOutlierRemoval(pfit, sigma_clip, niter=3, sigma=3.0)
                    model_pV, mask_V = sigma_clip_fit(p1, rPos, V_sig_c)
                    I_sig_c = I_sig_c - model_pI(rPos)
                    Q_sig_c = Q_sig_c - model_pQ(rPos)
                    U_sig_c = U_sig_c - model_pU(rPos)
                    V_sig_c = V_sig_c - model_pV(rPos)
                except:
                    pass
                self.I_sig_c = I_sig_c
                self.Q_sig_c = Q_sig_c
                self.U_sig_c = U_sig_c
                self.V_sig_c = V_sig_c

        if check:
            x = x[sel_i1i2]
            ax9 = fig.add_subplot(N_r_p, N_c_p, 9)
            ax9.plot(x, I_ref_c)
            ax10 = fig.add_subplot(N_r_p, N_c_p, 10)
            ax10.plot(x, Q_ref_c)
            ax11 = fig.add_subplot(N_r_p, N_c_p, 11)
            #ax11.plot(x, scale_IQ)
            ax11.plot(x, U_ref_c)
            ax12 = fig.add_subplot(N_r_p, N_c_p, 12)
            #ax12.plot(x, ang_inst)
            ax12.plot(x, V_ref_c)
            ax13 = fig.add_subplot(N_r_p, N_c_p, 13)
            ax13.plot(x, I_sig_c)
            ax14 = fig.add_subplot(N_r_p, N_c_p, 14)
            ax14.plot(x, Q_sig_c)
            ax15 = fig.add_subplot(N_r_p, N_c_p, 15)
            ax15.plot(x, U_sig_c)
            ax16 = fig.add_subplot(N_r_p, N_c_p, 16)
            ax16.plot(x, V_sig_c)
            ax17 = fig.add_subplot(N_r_p, N_c_p, 17)
            ax17.plot(x, parangle)
            plt.show()



def makeMapScan(kyDir, posFile0, infTemplate, fileNr, BeamList, ScanList, RA_c, Dec_c, scanDir, good_chan_file=None, tcalFile=None, gainRatioFile=None,baseFrac='all', RotAngle=Angle(0.*u.deg), numSkip=20, extent=1., grid=0.6,ChanList=[42000], interp2d='linear', interp=None, signQ=-1.0, signU=1.0, para_corr=False, check=False, checkMap=False, subBaseline=True, outbase=None, overwrite=False, baselineOrder=1, edgeFrac=[0.15,0.15], thres_max=None, editFile=None):

    ra_c = RA_c
    dec_c = Dec_c
    ext = extent 
    cdelt = grid/60. # arcmin
    Num = int(ext/cdelt+1)
    if Num %2 == 0: Num += 1
    ra_0 = ra_c + cdelt*(Num//2)
    dec_0 = dec_c - cdelt*(Num//2)
    ra_all = ra_0 - np.arange(Num) * cdelt 
    dec_all = dec_0 + np.arange(Num) * cdelt
    XI, YI = np.meshgrid(ra_all, dec_all)
    hdr = pf.Header({ 'SIMPLE':   True,
                      'BITPIX':    -32, 
                      'NAXIS':       2,
                      'NAXIS1':  Num,
                      'NAXIS2':  Num,
                      'CDELT1':  -cdelt,
                      'CRVAL1':  ra_c,
                      'CRPIX1':  (ra_c - ra_0)/(-cdelt) + 1.,
                      'CTYPE1': 'RA---CAR' ,
                      'CDELT2':   cdelt,
                      'CRPIX2':  (0.0 - dec_0) / cdelt + 1. ,
                      'CRVAL2':   0.0,
                      'CTYPE2': 'DEC--CAR' ,
                      'EXTEND':True
                    })
    hdr['HISTORY'] = 'Test ...'
    hdr['COMMENT'] = 'Test ...'

    if outbase == None: outbase = 'all'

    if tcalFile != None:
        tcals0 = np.fromfile(tcalFile, sep="\n")
        tcals = tcals0.reshape((tcals0.size//39, 39))

    if gainRatioFile != None:
        gainRatios = np.loadtxt(gainRatioFile, delimiter = ",")

    if good_chan_file != None:
        tmp = np.fromfile(good_chan_file, sep="\n")
        tmp = tmp.reshape((tmp.size//2, 2))
        good_flag = tmp[:,1]

    editDict = {}
    if editFile is not None:
        f_tmp = open(editFile, 'r')
        allLines_tmp = f_tmp.readlines()
        f_tmp.close()
        for term_tmp in allLines_tmp:
            x_tmp = term_tmp.split()
            if x_tmp[0] in editDict:
                if x_tmp[1] not in editDict[x_tmp[0]]:
                    editDict[x_tmp[0]][x_tmp[1]] = [float(x_tmp[2]), float(x_tmp[3])]
            else:
                editDict[x_tmp[0]] = {}
                editDict[x_tmp[0]][x_tmp[1]] = [float(x_tmp[2]), float(x_tmp[3])]

    print(editDict)


    dataDict = {} 
    for Chan in ChanList:
        try:
            print("Channel: ", Chan)
            if good_chan_file != None:
                if good_flag[Chan] < 0.5:
                    print("bad channel!")
                    continue

            outI_file = '%s_I_chan_%s.fits' % (outbase, str(Chan))
            outQ_file = '%s_Q_chan_%s.fits' % (outbase, str(Chan))
            outU_file = '%s_U_chan_%s.fits' % (outbase, str(Chan))
            outV_file = '%s_V_chan_%s.fits' % (outbase, str(Chan))
            outP_file = '%s_P_chan_%s.fits' % (outbase, str(Chan))

            if not overwrite:
                if os.path.exists(outI_file) and os.path.exists(outQ_file) and os.path.exists(outU_file) and os.path.exists(outV_file) and os.path.exists(outP_file): continue


            NN = 0
            for ii in ScanList:
                print("Scan: ",ii)
                posFile = posFile0.replace('ZZZ', str(ii))
                if not os.path.exists(posFile):
                    # Run for first time, calculate RA/DEC
                    kyFile = posFile.replace('csv', 'xlsx') 
                    track = Track(posFile, kyFile=kyFile)
                    track.runsave()
                else:
                    # Second time, read positions from the file
                    track = Track(posFile)
                    #track.plotRADec(output='ttt1.png')

                inf1 = infTemplate.replace('ZZZ', str(ii))

                for i in BeamList:
                    print("Beam: ",i)
                    BeamNr = "M%02d" % i
                    inf = inf1.replace('YYY', BeamNr[1:])

                    for kk in fileNr:
                        inf2 = inf.replace('XXX', kk)
                        key = BeamNr + 'S' + str(ii)
                        if key not in dataDict:
                            dataDict[key] = pf.getdata(inf2)
                        obs = ScanOne(dataDict[key], track.mjd, track.ra, track.dec, BeamNr, RotAngle, scanDir, numSkip=numSkip, Chan=Chan)
                        #obs.DisplayChannel()
                        #input('any key')
                        if tcalFile != None:
                            freq_t = tcals[:,0]
                            tcal_xx = tcals[:,2*i-1]
                            tcal_yy = tcals[:,2*i]
                            freq0 = obs.fghz*1.e3
                            tcal_interp_xx = interpolate.interp1d(freq_t, tcal_xx, kind='linear')
                            tcal_interp_yy = interpolate.interp1d(freq_t, tcal_yy, kind='linear')
                            Tcal = [tcal_interp_xx(freq0), tcal_interp_yy(freq0)]
                        else:
                            Tcal = [10., 10.]
                        obs.CalWithInjectedNoise(Tcal=Tcal, check=check, subBaseline=subBaseline, signQ=signQ, signU=signU, para_corr=para_corr, RotAngle=RotAngle, baseFrac=baseFrac, baselineOrder=baselineOrder, edgeFrac=edgeFrac)
                        #obs.CalWithInjectedNoise(Tcal=Tcal, check=True, subBaseline=subBaseline, signQ=signQ, signU=signU, para_corr=para_corr, RotAngle=RotAngle, baseFrac=baseFrac, baselineOrder=baselineOrder, edgeFrac=edgeFrac)

                        thres_max_flag = False
                        if thres_max is not None:
                            sel = obs.I_sig_c < thres_max
                            if obs.I_sig_c.size == obs.I_sig_c[sel].size: thres_max_flag = False
                            else: thres_max_flag = True

                        edit_flag = False
                        if str(ii) in editDict and str(i) in editDict[str(ii)]:
                            edit_flag = True

                        if scanDir == 'RA':
                            if len(obs.ra) == 0: continue
                            if thres_max_flag or edit_flag: tmp_x = obs.ra
                        else:
                            if len(obs.dec) == 0: continue
                            if thres_max_flag or edit_flag: tmp_x = obs.dec

                        if edit_flag:
                            sel1 = tmp_x < editDict[str(ii)][str(i)][0]
                            sel2 = tmp_x > editDict[str(ii)][str(i)][1]
                            if thres_max_flag: sel = sel & (sel1 | sel2)
                            else: sel = sel1 | sel2

                        #fig = plt.figure(figsize=(15,10))
                        #ax = fig.add_subplot(2,1,1)
                        #ax.plot(obs.dec, obs.I_sig_c,'o')
                        #ra_tmp = np.median(obs.ra)/15.
                        #ra_h_tmp = int(ra_tmp)
                        #ra_m_tmp = int((ra_tmp - ra_h_tmp) * 60.)
                        #ra_s_tmp = ra_tmp*3600. - ra_h_tmp * 3600. - ra_m_tmp * 60.
                        #ax.set_title(str(ra_h_tmp)+'h'+str(ra_m_tmp)+'m'+str(ra_s_tmp))

                        if thres_max_flag or edit_flag:

                            result = ndimage.median_filter(obs.I_sig_c[sel], size=20)
                            diff = obs.I_sig_c[sel] - result
                            mad_I = getMAD(diff)

                            result = ndimage.median_filter(obs.Q_sig_c[sel], size=20)
                            diff = obs.Q_sig_c[sel] - result
                            mad_Q = getMAD(diff)

                            result = ndimage.median_filter(obs.U_sig_c[sel], size=20)
                            diff = obs.U_sig_c[sel] - result
                            mad_U = getMAD(diff)

                            result = ndimage.median_filter(obs.V_sig_c[sel], size=20)
                            diff = obs.V_sig_c[sel] - result
                            mad_V = getMAD(diff)

                            f_y1 = interpolate.interp1d(tmp_x[sel], obs.I_sig_c[sel], kind='linear')
                            f_y2 = interpolate.interp1d(tmp_x[sel], obs.Q_sig_c[sel], kind='linear')
                            f_y3 = interpolate.interp1d(tmp_x[sel], obs.U_sig_c[sel], kind='linear')
                            f_y4 = interpolate.interp1d(tmp_x[sel], obs.V_sig_c[sel], kind='linear')

                            #f_y1 = interpolate.interp1d(tmp_x[sel], obs.I_sig_c[sel], kind='nearest')
                            #f_y2 = interpolate.interp1d(tmp_x[sel], obs.Q_sig_c[sel], kind='nearest')
                            #f_y3 = interpolate.interp1d(tmp_x[sel], obs.U_sig_c[sel], kind='nearest')
                            #f_y4 = interpolate.interp1d(tmp_x[sel], obs.V_sig_c[sel], kind='nearest')

                            #f_y1 = interpolate.InterpolatedUnivariateSpline(tmp_x[sel], obs.I_sig_c[sel])
                            #f_y2 = interpolate.InterpolatedUnivariateSpline(tmp_x[sel], obs.Q_sig_c[sel])
                            #f_y3 = interpolate.InterpolatedUnivariateSpline(tmp_x[sel], obs.U_sig_c[sel])
                            #f_y4 = interpolate.InterpolatedUnivariateSpline(tmp_x[sel], obs.V_sig_c[sel])

                            for jjj in range(obs.I_sig_c.size):
                                if thres_max_flag:
                                    if obs.I_sig_c[jjj] >= thres_max:
                                        obs.I_sig_c[jjj] = f_y1(tmp_x[jjj]) + np.random.normal(loc=0., scale=mad_I) 
                                        obs.Q_sig_c[jjj] = f_y2(tmp_x[jjj]) + np.random.normal(loc=0., scale=mad_Q)
                                        obs.U_sig_c[jjj] = f_y3(tmp_x[jjj]) + np.random.normal(loc=0., scale=mad_U)
                                        obs.V_sig_c[jjj] = f_y4(tmp_x[jjj]) + np.random.normal(loc=0., scale=mad_V)
                                if edit_flag:
                                    if tmp_x[jjj] >= editDict[str(ii)][str(i)][0] and tmp_x[jjj]<=editDict[str(ii)][str(i)][1]:
                                        obs.I_sig_c[jjj] = f_y1(tmp_x[jjj]) + np.random.normal(loc=0., scale=mad_I)
                                        obs.Q_sig_c[jjj] = f_y2(tmp_x[jjj]) + np.random.normal(loc=0., scale=mad_Q)
                                        obs.U_sig_c[jjj] = f_y3(tmp_x[jjj]) + np.random.normal(loc=0., scale=mad_U)
                                        obs.V_sig_c[jjj] = f_y4(tmp_x[jjj]) + np.random.normal(loc=0., scale=mad_V)

                        #ax.plot(obs.dec, obs.I_sig_c,'ro')
                        #if edit_flag:
                        #    ax.plot(obs.dec, f_y1(obs.dec))
                        #    ax1 = fig.add_subplot(2,1,2)
                        #    ax1.plot(obs.dec, obs.I_sig_c,'ro')
                        #    ax1.plot(obs.dec, f_y1(obs.dec))
                        #plt.show()
                        #input('any key')

                        if interp == 'expsinc':
                            if scanDir == 'RA':
                                I_tt = expsinc(obs.ra, obs.I_sig_c, ra_all, DIAM, obs.fghz)
                                Q_tt = expsinc(obs.ra, obs.Q_sig_c, ra_all, DIAM, obs.fghz)
                                U_tt = expsinc(obs.ra, obs.U_sig_c, ra_all, DIAM, obs.fghz)
                                V_tt = expsinc(obs.ra, obs.V_sig_c, ra_all, DIAM, obs.fghz)
                                ra_tt = ra_all * 1.
                                f_dec = interpolate.interp1d(obs.ra, obs.dec, fill_value=-999999., bounds_error=False)
                                dec_tt = f_dec(ra_tt)

                                I_tt = I_tt[dec_tt > -10000.]
                                Q_tt = Q_tt[dec_tt > -10000.]
                                U_tt = U_tt[dec_tt > -10000.]
                                V_tt = V_tt[dec_tt > -10000.]
                                ra_tt = ra_tt[dec_tt > -10000.]
                                dec_tt = dec_tt[dec_tt > -10000.]

                            else:
                                I_tt = expsinc(obs.dec, obs.I_sig_c, dec_all, DIAM, obs.fghz)
                                Q_tt = expsinc(obs.dec, obs.Q_sig_c, dec_all, DIAM, obs.fghz)
                                U_tt = expsinc(obs.dec, obs.U_sig_c, dec_all, DIAM, obs.fghz)
                                V_tt = expsinc(obs.dec, obs.V_sig_c, dec_all, DIAM, obs.fghz)
                                dec_tt = dec_all * 1.
                                f_ra = interpolate.interp1d(obs.dec, obs.ra, fill_value=-999999., bounds_error=False)
                                ra_tt = f_ra(dec_tt)

                                I_tt = I_tt[ra_tt > -10000.]
                                Q_tt = Q_tt[ra_tt > -10000.]
                                U_tt = U_tt[ra_tt > -10000.]
                                V_tt = V_tt[ra_tt > -10000.]
                                dec_tt = dec_tt[ra_tt > -10000.]
                                ra_tt = ra_tt[ra_tt > -10000.]

                        else:
                            I_tt = obs.I_sig_c * 1.
                            Q_tt = obs.Q_sig_c * 1.
                            U_tt = obs.U_sig_c * 1.
                            V_tt = obs.V_sig_c * 1.
                            ra_tt = obs.ra * 1.
                            dec_tt = obs.dec * 1.


                        if gainRatioFile != None:
                            f_tmp = gainRatios[i-1][0] + gainRatios[i-1][1] * obs.fghz
                            I_tt *= f_tmp
                            Q_tt *= f_tmp
                            U_tt *= f_tmp
                            V_tt *= f_tmp

                        if NN == 0:
                            ra0 = ra_tt
                            dec0 = dec_tt
                            I0 = I_tt
                            Q0 = Q_tt
                            U0 = U_tt
                            V0 = V_tt
                        else:
                            ra0 = np.concatenate((ra0, ra_tt))
                            dec0 = np.concatenate((dec0, dec_tt))
                            I0 = np.concatenate((I0, I_tt))
                            Q0 = np.concatenate((Q0, Q_tt))
                            U0 = np.concatenate((U0, U_tt))
                            V0 = np.concatenate((V0, V_tt))
                        NN += 1
                        #print(ra0, dec0)
                        #print(I0.size)
                        #print(I0)
                        #input('any key')

            #print(ra0, dec0)
            #print(XI, YI)
            #print(I0)
            
            #2.29_plt
            # print('Track_ScanDec_beam')
            # plt.scatter(ra0,dec0,c=I0)
            # plt.savefig('Track_ScanDec_beam.png')

            I_all = griddata((ra0, dec0), I0, (XI, YI), method=interp2d, fill_value=0.)
            Q_all = griddata((ra0, dec0), Q0, (XI, YI), method=interp2d, fill_value=0.)
            U_all = griddata((ra0, dec0), U0, (XI, YI), method=interp2d, fill_value=0.)
            V_all = griddata((ra0, dec0), V0, (XI, YI), method=interp2d, fill_value=0.)


            pf.writeto(outI_file, I_all, hdr, overwrite=True)
            pf.writeto(outQ_file, Q_all, hdr, overwrite=True)
            pf.writeto(outU_file, U_all, hdr, overwrite=True)
            pf.writeto(outV_file, V_all, hdr, overwrite=True)
            pdata = np.sqrt(Q_all*Q_all + U_all*U_all)
            pf.writeto(outP_file, pdata, hdr, overwrite=True)
            del I_all
            del Q_all
            del U_all
            del V_all
            del pdata
            gc.collect()

            if checkMap:
                fig = plt.figure()
                ax1 = fig.add_subplot(2,2,1)
                plotImage(ax1, outI_file, glvisible=False, gbvisible=True, gltickvisible=False, gbtickvisible=True)
                ax2 = fig.add_subplot(2,2,2)
                plotImage(ax2, outQ_file, glvisible=False, gbvisible=False, gltickvisible=False, gbtickvisible=False)
                ax3 = fig.add_subplot(2,2,3)
                plotImage(ax3, outU_file, glvisible=True, gbvisible=True, gltickvisible=True, gbtickvisible=True)
                ax4 = fig.add_subplot(2,2,4)
                plotImage(ax4, outV_file, glvisible=True, gbvisible=False, gltickvisible=True, gbtickvisible=False)
                plt.tight_layout()
                #plt.show()
                fig.savefig('%s_IQUV_chan_%s.png' % (outbase, str(Chan)))
                plt.close()
        except Exception as e:
                print(f"Error processing scan {ii} on channel {Chan}: {e}")
                continue


def main():

    home = '/share/home/whjing/fast/'
    kyDir = home + 'KY/'
    
    posFile = home + 'file_redo-day-3/G182_RZZZ.csv'
    infTemplate = home + "file_redo-day-3/G182_RZZZ-MYYY_000XXX.fits"
    RA_c = 84.187500
    Dec_c = 24.675000
    scanDir = 'RA' #'RA'
    
    fileNr = ['1']
    BeamList = range(1,20)
    ScanList = range(1,6)
    
    Nchan = 65536

    args = sys.argv[1:]
    if len(args) == 1:#run all image in hpc
        Nchan_N = int(args[0])
        Nchan = 1100
        nchan_o = Nchan*(Nchan_N-1)
        nchan_f = Nchan*Nchan_N
        if nchan_f > 65536:
            nchan_f = 65536
        ChanList = np.arange(nchan_o,nchan_f)  
    else:
        ChanList = [40000] #test one image


    numSkip = 20 
    extent = 5.
    grid = 0.6 # arcmin
    outbase = home + 'image_redo-day-3/G182_RA'#test_name_type
    ang = -23.4 #Ra:23.4 or -23.4; Dec:-6.6(from obsplan)
    interp='2dlinear'
    signQ = -1.
    signU = 1.
    
    good_chan_file = '../code_i/good-chans-rms_2.txt'
    # good_chan_file = home + 'output/good-chan_modified.dat'#'3C380/good_chan.dat'
    tcalFile = home + 'cal/low_cal_my.dat'
    gainRatioFile = home + '3c138_0511/gain_ratio_R.txt'


    para_corr = False
    check = False
    checkMap = False
    # baseFrac='double-edge'
    baseFrac = 'all'
    subBaseline = True
    interp2d = 'linear'
    # interp2d = 'expsinc'
    interp = 'none'
    overwrite = True
    baselineOrder = 1
    edgeFrac = [0.1, 0.1]
    thres_max = None 
    editFile = None
    print(subBaseline)
    

    makeMapScan(kyDir, posFile, infTemplate, fileNr, BeamList, ScanList, RA_c, Dec_c, scanDir, good_chan_file=good_chan_file, tcalFile=tcalFile, gainRatioFile=gainRatioFile,baseFrac=baseFrac,RotAngle=Angle(ang*u.deg), numSkip=numSkip, extent=extent, grid=grid, ChanList=ChanList, interp2d=interp2d, interp=interp, signQ=signQ, signU=signU, para_corr=para_corr, check=check, checkMap=checkMap, subBaseline=subBaseline, outbase=outbase, overwrite=overwrite,baselineOrder=baselineOrder,edgeFrac=edgeFrac, thres_max=thres_max, editFile=editFile)

if __name__ == "__main__":
    main()
